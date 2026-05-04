#!/usr/bin/env python3
"""
Generate confusion matrix figures comparing LLM grader vs Human grader.

For every CombinedHumanGradingAndLLMGrading.xlsx the script produces, inside a
  confusion_matrices/  subfolder next to the xlsx:
    • confusion_matrices.png  — combined 2×4 grid (1 global + 7 element types)
    • global.png, state.png, transition.png, composite_state.png,
      guard.png, action.png, history_state.png, region.png  — 8 individual figures

Additionally, an aggregated analysis across ALL xlsx files is written to:
    global_analysis/confusion_matrices/   (relative to this script's directory)

Usage:
  python generate_agreement_figures.py            # all files
  python generate_agreement_figures.py --stage 2  # 2-stage files only
  python generate_agreement_figures.py --stage 1  # 1-stage files only
  python generate_agreement_figures.py --date 2026-03-16
"""

import argparse
import glob
import os
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SCORE_LABELS = ["0", "0.5", "1"]
SCORE_TO_IDX = {0.0: 0, 0.5: 1, 1.0: 2}

# Ordered list of (display title, filename stem, optional type-filter string)
SCOPES: list[tuple[str, str, str | None]] = [
    ("Global", "global", None),
    ("State", "state", "State"),
    ("Transition", "transition", "Transition"),
    ("Composite State", "composite_state", "Composite State"),
    ("Guard", "guard", "Guard"),
    ("Action", "action", "Action"),
    ("History State", "history_state", "History State"),
    ("Region", "region", "Region"),
]

BASE_DIR = Path(__file__).parent

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def find_xlsx_files(
    base_dir: Path,
    stage_filter: str | None = None,
    date_filter: str | None = None,
) -> list[Path]:
    """Return all CombinedHumanGradingAndLLMGrading.xlsx files (no temp ~$ files)."""
    pattern = str(base_dir / "**" / "*CombinedHumanGradingAndLLMGrading.xlsx")
    files = [
        Path(f)
        for f in glob.glob(pattern, recursive=True)
        if not os.path.basename(f).startswith("~$")
    ]
    if stage_filter:
        tag = f"{stage_filter} stage"
        files = [f for f in files if tag in str(f)]
    if date_filter:
        files = [f for f in files if date_filter in str(f)]
    return sorted(files)


def _find_kappa_sheet(wb: openpyxl.Workbook) -> str | None:
    """Return the name of the weighted Cohen's kappa sheet, regardless of exact capitalisation."""
    for name in wb.sheetnames:
        lower = name.lower()
        if "kappa" in lower or ("cohen" in lower and "weight" in lower):
            return name
    return None


def load_grading_data(xlsx_path: Path) -> pd.DataFrame | None:
    """
    Load Human and LLM grades from the weighted Cohen's kappa sheet.
    Returns a DataFrame with columns: type, element, human, llm, is_additional.

    Regular rows have is_additional=False and human/llm are 0/0.5/1 grading scores.
    'additional elements' rows have is_additional=True and human/llm are raw *counts*
    of false-positive elements; they are handled separately in build_confusion_matrix.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"  [WARN] Could not open {xlsx_path.name}: {e}")
        return None

    sheet_name = _find_kappa_sheet(wb)
    if sheet_name is None:
        print(f"  [WARN] No weighted Cohen's kappa sheet found in {xlsx_path.name}")
        return None

    ws = wb[sheet_name]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        type_, element, human, llm = row[0], row[1], row[2], row[3]
        if type_ is None or element is None:
            break
        if human is None or llm is None:
            continue
        is_additional = str(element).strip().lower() == "additional elements"
        try:
            records.append(
                {
                    "type": str(type_).strip(),
                    "element": str(element).strip(),
                    "human": float(human),
                    "llm": float(llm),
                    "is_additional": is_additional,
                }
            )
        except (ValueError, TypeError):
            continue

    return pd.DataFrame(records) if records else None


def load_all(files: list[Path]) -> pd.DataFrame:
    """Load and concatenate grading data from all xlsx files."""
    frames = []
    for f in files:
        df = load_grading_data(f)
        if df is not None and len(df) > 0:
            df["source"] = f.name
            frames.append(df)
            print(f"  Loaded {len(df):3d} rows  ← {f.name[:75]}")
    if not frames:
        raise ValueError("No valid data found in any xlsx file.")
    return pd.concat(frames, ignore_index=True)


# ---------------------------------------------------------------------------
# Confusion matrix helpers
# ---------------------------------------------------------------------------


def build_confusion_matrix(
    df: pd.DataFrame, type_filter: str | None = None
) -> np.ndarray:
    """Build a 3×3 count matrix (rows = Human score, cols = LLM score).

    Regular rows (is_additional=False) contribute via their 0/0.5/1 scores.

    'Additional elements' rows carry raw counts of false-positive elements and
    are distributed across cells using the same MAX/MIN logic as the Excel sheet:
      - MIN(human, llm)          → [human=1, llm=1]  both graders agreed on extras
      - MAX(llm - human, 0)      → [human=0, llm=1]  LLM hallucinated extra elements
      - MAX(human - llm, 0)      → [human=1, llm=0]  human found extras LLM missed
    """
    if type_filter:
        df = df[df["type"] == type_filter]

    regular = df[~df["is_additional"]]
    additional = df[df["is_additional"]]

    matrix = np.zeros((3, 3), dtype=int)

    for _, row in regular.iterrows():
        h = SCORE_TO_IDX.get(row["human"])
        l = SCORE_TO_IDX.get(row["llm"])
        if h is not None and l is not None:
            matrix[h, l] += 1

    for _, row in additional.iterrows():
        h_count = int(row["human"])
        l_count = int(row["llm"])
        if h_count == 0 and l_count == 0:
            matrix[0, 0] += 1  # both agree there are no extra elements
        else:
            matrix[2, 2] += min(h_count, l_count)  # both agreed on extras
            matrix[0, 2] += max(l_count - h_count, 0)  # LLM found more extras
            matrix[2, 0] += max(h_count - l_count, 0)  # human found more extras

    return matrix


def row_normalize(matrix: np.ndarray) -> np.ndarray:
    """Normalize each row to [0, 1]; rows that sum to zero become NaN."""
    row_sums = matrix.sum(axis=1, keepdims=True).astype(float)
    row_sums[row_sums == 0] = np.nan
    return matrix / row_sums


# ---------------------------------------------------------------------------
# Core plotting primitive
# ---------------------------------------------------------------------------


def plot_confusion_heatmap(
    ax: plt.Axes,
    matrix: np.ndarray,
    title: str,
    show_ylabel: bool = True,
    show_xlabel: bool = True,
) -> None:
    """Draw a single row-normalized confusion-matrix heatmap onto ax."""
    norm = row_normalize(matrix)

    annots = np.empty_like(norm, dtype=object)
    for i in range(3):
        for j in range(3):
            pct = norm[i, j]
            cnt = matrix[i, j]
            annots[i, j] = "—" if np.isnan(pct) else f"{pct:.0%}\n({cnt})"

    sns.heatmap(
        np.where(np.isnan(norm), 0, norm),
        annot=annots,
        fmt="",
        cmap="Blues",
        xticklabels=SCORE_LABELS,
        yticklabels=SCORE_LABELS,
        ax=ax,
        vmin=0,
        vmax=1,
        cbar=False,
        linewidths=0.5,
        linecolor="white",
    )
    for i in range(3):
        ax.add_patch(
            plt.Rectangle(
                (i, i), 1, 1, fill=False, edgecolor="black", lw=1.5, clip_on=False
            )
        )

    n_total = int(matrix.sum())
    n_agree = int(np.trace(matrix))
    agree_pct = n_agree / n_total if n_total > 0 else 0
    ax.set_title(f"{title}\n(agree {agree_pct:.0%}, n={n_total})", fontsize=9, pad=4)
    ax.set_xlabel("LLM Score" if show_xlabel else "", fontsize=8)
    ax.set_ylabel("Human Score" if show_ylabel else "", fontsize=8)
    ax.tick_params(labelsize=8)


# ---------------------------------------------------------------------------
# Figure A: Combined 2×4 grid (unchanged from original)
# ---------------------------------------------------------------------------


def save_confusion_grid(
    df: pd.DataFrame, out_path: Path, title_prefix: str = ""
) -> None:
    """Save the 2×4 combined confusion-matrix grid to out_path."""
    fig, axes = plt.subplots(2, 4, figsize=(14, 7))
    axes = axes.flatten()

    for idx, (title, _, type_filter) in enumerate(SCOPES):
        matrix = build_confusion_matrix(df, type_filter=type_filter)
        plot_confusion_heatmap(
            axes[idx],
            matrix,
            title,
            show_ylabel=(idx % 4 == 0),
            show_xlabel=True,
        )

    # No unused axes: 8 scopes fill the 2×4 grid exactly.
    sup = f"{title_prefix} — " if title_prefix else ""
    fig.suptitle(
        f"{sup}LLM vs Human Grader: Row-Normalized Confusion Matrices\n"
        "(each row shows how the LLM scored items the human scored 0 / 0.5 / 1)",
        fontsize=10,
        y=1.01,
    )
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved → {out_path.relative_to(BASE_DIR)}")


# ---------------------------------------------------------------------------
# Figure B: 8 individual confusion matrices
# ---------------------------------------------------------------------------


def save_individual_confusion_matrices(
    df: pd.DataFrame, out_dir: Path, title_prefix: str = "", file_prefix: str = ""
) -> None:
    """Save one figure per scope (global.png, state.png, …) into out_dir."""
    out_dir.mkdir(parents=True, exist_ok=True)
    sup = f"{title_prefix} — " if title_prefix else ""

    for title, stem, type_filter in SCOPES:
        matrix = build_confusion_matrix(df, type_filter=type_filter)
        fig, ax = plt.subplots(figsize=(4, 3.5))
        plot_confusion_heatmap(ax, matrix, title, show_ylabel=True, show_xlabel=True)
        fig.suptitle(
            f"{sup}LLM vs Human: {title}",
            fontsize=9,
            y=1.02,
        )
        fig.tight_layout()
        fname = f"{file_prefix}_{stem}.png" if file_prefix else f"{stem}.png"
        out_path = out_dir / fname
        fig.savefig(out_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        print(f"  Saved → {out_path.relative_to(BASE_DIR)}")


# ---------------------------------------------------------------------------
# High-level: write all confusion matrix outputs into a given directory
# ---------------------------------------------------------------------------


def generate_confusion_outputs(
    df: pd.DataFrame, cm_dir: Path, title_prefix: str = "", file_prefix: str = ""
) -> None:
    """
    Write into cm_dir:
      [file_prefix_]confusion_matrices.png   (combined 2×4 grid)
      [file_prefix_]global.png, state.png, … (8 individual figures)
    """
    cm_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n  → {cm_dir.relative_to(BASE_DIR)}  ({len(df)} rows)")
    grid_name = (
        f"{file_prefix}_confusion_matrices.png"
        if file_prefix
        else "confusion_matrices.png"
    )
    save_confusion_grid(df, cm_dir / grid_name, title_prefix)
    save_individual_confusion_matrices(
        df, cm_dir, title_prefix, file_prefix=file_prefix
    )


# ---------------------------------------------------------------------------
# Label helper
# ---------------------------------------------------------------------------


def label_from_path(xlsx_path: Path) -> str:
    """
    Derive a human-readable label from the xlsx path, e.g.:
      Printer — 2 stage — 2026-03-16
    """
    parts = xlsx_path.parts
    # Walk up from the file to find the date folder, stage folder, and project folder
    # Expected layout: …/<Project>/Grading/<N stage>/<Date>/<file>.xlsx
    date_dir = xlsx_path.parent.name  # e.g. "2026-03-16"
    stage_dir = xlsx_path.parent.parent.name  # e.g. "2 stage"
    # Project is two levels above "Grading"
    try:
        grading_idx = parts.index("Grading")
        project = parts[grading_idx - 1]
    except (ValueError, IndexError):
        project = xlsx_path.parent.parent.parent.name
    return f"{project} — {stage_dir} — {date_dir}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate confusion matrix figures for LLM vs Human grader."
    )
    parser.add_argument(
        "--stage",
        choices=["1", "2"],
        default=None,
        help="Restrict to '1 stage' or '2 stage' files. Default: all.",
    )
    parser.add_argument(
        "--date",
        default=None,
        help="Restrict to files whose path contains this date string (e.g. 2026-03-16).",
    )
    args = parser.parse_args()

    print(
        f"Searching for xlsx files (stage={args.stage or 'any'}, date={args.date or 'any'}) …"
    )
    files = find_xlsx_files(BASE_DIR, stage_filter=args.stage, date_filter=args.date)
    print(f"Found {len(files)} file(s).\n")

    if not files:
        print("No files matched. Exiting.")
        return

    # ------------------------------------------------------------------
    # 1. Per-file: output into <date_dir>/confusion_matrices/
    # ------------------------------------------------------------------
    print("=" * 60)
    print("Per-file confusion matrices")
    print("=" * 60)
    all_frames: list[pd.DataFrame] = []
    for xlsx_path in files:
        df = load_grading_data(xlsx_path)
        if df is None or len(df) == 0:
            print(f"  [SKIP] {xlsx_path.name}")
            continue
        all_frames.append(df)
        cm_dir = xlsx_path.parent / "confusion_matrices"
        label = label_from_path(xlsx_path)
        generate_confusion_outputs(
            df, cm_dir, title_prefix=label, file_prefix=xlsx_path.stem
        )

    if not all_frames:
        print("No valid data found. Exiting.")
        return

    # ------------------------------------------------------------------
    # 2. Aggregate: output into global_analysis/confusion_matrices/
    # ------------------------------------------------------------------
    print()
    print("=" * 60)
    print("Aggregated confusion matrices (all files)")
    print("=" * 60)
    df_all = pd.concat(all_frames, ignore_index=True)
    global_cm_dir = BASE_DIR / "global_analysis" / "confusion_matrices"
    generate_confusion_outputs(
        df_all, global_cm_dir, title_prefix="All Examples", file_prefix="all_examples"
    )

    print("\nDone.")


if __name__ == "__main__":
    main()
