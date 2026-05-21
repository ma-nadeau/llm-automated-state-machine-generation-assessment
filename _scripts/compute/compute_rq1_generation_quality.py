#!/usr/bin/env python3
"""
RQ1: Human assessment of Claude-generated state machines.

Reads the human-grading section (top of the Metrics sheet, rows 3-10) from every
*CombinedHumanGradingVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration.xlsx
workbook (all stages/shots), pools raw N / TP / FP / FN counts across all examples
per (approach, element type), and computes micro-averaged Precision / Recall / F1
from the pooled totals.

Outputs (written to _Figures/data/rq1/):
  rq1_generation_quality_summary.csv    — pooled counts + metrics per approach and element type
  rq1_generation_quality_by_example.csv — one row per (approach, project, element type) with raw counts

Usage:
  python compute_rq1_generation_quality.py                         # writes CSVs to default out-dir
  python compute_rq1_generation_quality.py --out-dir /custom/path  # override output directory
  python compute_rq1_generation_quality.py --print                 # print Markdown table to stdout only
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl

from _xlsx_utils import (
    BASE_DIR,
    TYPES,
    MetricRow,
    approach_from_path,
    as_float,
    fmt,
    iter_xlsx,
    md_table,
    project_from_path,
    write_csv,
)

DEFAULT_OUT_DIR = BASE_DIR / "_Figures" / "data" / "rq1"


def _compute_metrics(
    tp: float, fp: float, fn: float
) -> tuple[float | None, float | None, float | None]:
    precision = tp / (tp + fp) if (tp + fp) > 0 else None
    recall = tp / (tp + fn) if (tp + fn) > 0 else None
    if precision is not None and recall is not None and (precision + recall) > 0:
        f1 = 2 * precision * recall / (precision + recall)
    else:
        f1 = None
    return precision, recall, f1


def load_human_metric_rows(root: Path) -> list[MetricRow]:
    files = sorted(iter_xlsx(root.glob(
        "*/Grading/*/*/"
        "*_CombinedHumanGradingVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration.xlsx"
    )))
    rows: list[MetricRow] = []
    for path in files:
        wb = openpyxl.load_workbook(path, data_only=True)
        metrics_name = next((s for s in wb.sheetnames if "metric" in s.lower()), None)
        if not metrics_name:
            wb.close()
            continue
        ws = wb[metrics_name]
        # Human section: header at row 2, data at rows 3-9, overall at row 10
        for rn in range(3, 11):
            raw_type = ws.cell(rn, 1).value
            if not raw_type:
                continue
            if str(raw_type).strip().lower() == "overall score":
                continue
            label = str(raw_type).strip().replace("Composite state", "Composite State")
            if label not in ELEMENT_TYPES:
                continue
            rows.append(
                MetricRow(
                    approach=approach_from_path(path),
                    project=project_from_path(path),
                    model_element=label,
                    n=as_float(ws.cell(rn, 2).value),
                    tp=as_float(ws.cell(rn, 3).value),
                    fp=as_float(ws.cell(rn, 4).value),
                    fn=as_float(ws.cell(rn, 5).value),
                )
            )
        wb.close()
    return rows


ELEMENT_TYPES = sorted(t for t in TYPES if t != "Overall Score")


def summarize_rq1(root: Path) -> list[dict[str, object]]:
    metric_rows = load_human_metric_rows(root)
    grouped: dict[tuple[str, str], list[MetricRow]] = defaultdict(list)
    for row in metric_rows:
        grouped[(row.approach, row.model_element)].append(row)

    rows = []
    for approach in ("one-stage", "two-stage (3 examples)", "two-stage (6 examples)"):
        total_n = total_tp = total_fp = total_fn = 0.0
        approach_projects: set[str] = set()
        has_data = False

        for model_element in ELEMENT_TYPES:
            values = grouped.get((approach, model_element), [])
            if not values:
                continue
            has_data = True
            n = sum(v.n or 0 for v in values)
            tp = sum(v.tp or 0 for v in values)
            fp = sum(v.fp or 0 for v in values)
            fn = sum(v.fn or 0 for v in values)
            precision, recall, f1 = _compute_metrics(tp, fp, fn)
            rows.append(
                {
                    "approach": approach,
                    "model_element": model_element,
                    "examples": len({v.project for v in values}),
                    "n": n,
                    "tp": tp,
                    "fp": fp,
                    "fn": fn,
                    "precision": round(precision, 3) if precision is not None else None,
                    "recall": round(recall, 3) if recall is not None else None,
                    "f1": round(f1, 3) if f1 is not None else None,
                }
            )
            total_n += n
            total_tp += tp
            total_fp += fp
            total_fn += fn
            approach_projects.update(v.project for v in values)

        if has_data:
            precision, recall, f1 = _compute_metrics(total_tp, total_fp, total_fn)
            rows.append(
                {
                    "approach": approach,
                    "model_element": "Overall",
                    "examples": len(approach_projects),
                    "n": total_n,
                    "tp": total_tp,
                    "fp": total_fp,
                    "fn": total_fn,
                    "precision": round(precision, 3) if precision is not None else None,
                    "recall": round(recall, 3) if recall is not None else None,
                    "f1": round(f1, 3) if f1 is not None else None,
                }
            )
    return rows


def rq1_markdown(root: Path) -> str:
    table_rows = [
        [
            str(r["approach"]),
            str(r["model_element"]),
            str(r["examples"]),
            fmt(r["precision"]),
            fmt(r["recall"]),
            fmt(r["f1"]),
        ]
        for r in summarize_rq1(root)
    ]
    return md_table(
        ["Approach", "Model element", "Examples", "Precision", "Recall", "F1"],
        table_rows,
    )


def write_rq1(root: Path, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    write_csv(
        out_dir / "rq1_generation_quality_summary.csv",
        ["approach", "model_element", "examples", "n", "tp", "fp", "fn", "precision", "recall", "f1"],
        summarize_rq1(root),
    )
    metric_rows = load_human_metric_rows(root)
    write_csv(
        out_dir / "rq1_generation_quality_by_example.csv",
        ["approach", "project", "model_element", "n", "tp", "fp", "fn"],
        [
            {
                "approach": r.approach,
                "project": r.project,
                "model_element": r.model_element,
                "n": r.n,
                "tp": r.tp,
                "fp": r.fp,
                "fn": r.fn,
            }
            for r in metric_rows
        ],
    )
    print(f"RQ1 CSVs written to: {out_dir}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Summarize RQ1 generation quality.")
    parser.add_argument("--root", type=Path, default=BASE_DIR)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--print", action="store_true", help="Print Markdown table to stdout.")
    args = parser.parse_args()

    if args.print:
        print("## RQ1: Claude Generation Quality From Human Assessment\n")
        print(rq1_markdown(args.root))
    else:
        write_rq1(args.root, args.out_dir)


if __name__ == "__main__":
    main()
