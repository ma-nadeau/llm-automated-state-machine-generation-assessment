#!/usr/bin/env python3
"""
Aggregate CombinedHumanGrading vs LLM AutoGrading metrics for the 2-stage 6-examples setup.

Reads the Metrics sheet of every CombinedHumanGradingVs<LLM>AutoGrading workbook and
extracts two grading sections per file:

  Rows 1-10  – Human grading (human-annotated diagrams evaluated by a human grader).
               Identical across the 3 LLM files for the same example, so it is read
               only once per example.

  Rows 12-21 – LLM auto-grader section (Claude4.5Sonnet-generated diagrams evaluated
               by the LLM named in the filename).


For each section, N, TP, FP, and  FN are taken from the cached Metrics sheet values.

Aggregates across all 9 examples per (Grader, ElementType), then across all element
types for an "Overall" row per grader.

Output: a single merged CSV saved to _Figures/data/rq2/
  rq2_2stage_6examples_Metrics_CombinedHumanVsLLM.csv
Columns: Grader, ElementType, TotalN, TotalTP, TotalFP, TotalFN, TotalElements, Precision, Recall, F1
  TotalElements = TotalN + TotalFP (all expected + all spurious generated elements)
"""

import csv
import glob
import os
import re
from collections import defaultdict

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
OUTPUT_DIR = os.path.join(BASE_DIR, "_Figures", "data", "rq2")
OUTPUT_FILE = "rq2_2stage_6examples_Metrics_CombinedHumanVsLLM.csv"

TARGET_TYPES = [
    "Action",
    "Composite State",
    "Guard",
    "History State",
    "Region",
    "State",
    "Transition",
]

TYPE_ALIASES = {
    "action": "Action",
    "composite state": "Composite State",
    "composite states": "Composite State",
    "guard": "Guard",
    "history state": "History State",
    "history states": "History State",
    "region": "Region",
    "state": "State",
    "states": "State",
    "transition": "Transition",
    "transitions": "Transition",
}

HUMAN_GRADER_LABEL = "Human"


def normalize_type(raw):
    if raw is None:
        return None
    return TYPE_ALIASES.get(str(raw).strip().lower())


def safe_float(val):
    if val is None or val == "Not Available":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def compute_metrics(tp, fp, fn):
    precision = tp / (tp + fp) if (tp + fp) > 0 else None
    recall = tp / (tp + fn) if (tp + fn) > 0 else None
    if precision is not None and recall is not None and (precision + recall) > 0:
        f1 = 2 * precision * recall / (precision + recall)
    else:
        f1 = None
    return precision, recall, f1


def fmt(val, decimals=6):
    return round(val, decimals) if val is not None else "N/A"


def read_section(ws, start_row, end_row):
    """
    Read target element type rows between start_row and end_row (exclusive of header
    and 'Overall Score').  Returns dict: norm_type -> (N, TP, FP, FN_effective).
    """
    result = {}
    for rn in range(start_row, end_row):
        raw_type = ws.cell(rn, 1).value
        if not raw_type:
            continue
        if str(raw_type).strip().lower() == "overall score":
            continue
        norm_type = normalize_type(raw_type)
        if norm_type is None:
            continue
        n = safe_float(ws.cell(rn, 2).value) or 0.0
        tp = safe_float(ws.cell(rn, 3).value) or 0.0
        fp = safe_float(ws.cell(rn, 4).value) or 0.0
        fn = safe_float(ws.cell(rn, 5).value) or 0.0 
        result[norm_type] = (n, tp, fp, fn)
    return result


def accumulate(agg, section_data):
    for norm_type, (n, tp, fp, fn) in section_data.items():
        agg[norm_type]["N"] += n
        agg[norm_type]["TP"] += tp
        agg[norm_type]["FP"] += fp
        agg[norm_type]["FN"] += fn


def main():
    pattern = os.path.join(
        BASE_DIR,
        "*/Grading/2 stage/6-examples/"
        "*_CombinedHumanGradingVs*AutoGrading_Claude4.5SonnetGeneration.xlsx",
    )
    files = sorted(
        f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~")
    )

    if not files:
        print("ERROR: No matching files found.")
        return

    print(f"Found {len(files)} files.\n")

    # grader -> element_type -> cumulative {N, TP, FP, FN}
    aggregated = defaultdict(
        lambda: defaultdict(lambda: {"N": 0.0, "TP": 0.0, "FP": 0.0, "FN": 0.0})
    )

    # Track which examples have already contributed to the Human aggregate
    human_examples_seen = set()

    for filepath in files:
        bn = os.path.basename(filepath)
        m = re.search(r"CombinedHumanGradingVs(.+?)AutoGrading", bn)
        if not m:
            print(f"  WARNING: Cannot parse LLM name from: {bn}")
            continue
        llm = m.group(1)
        example = os.path.basename(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(filepath))))
        )

        wb = openpyxl.load_workbook(filepath, data_only=True)
        metrics_name = next((s for s in wb.sheetnames if "metric" in s.lower()), None)
        if not metrics_name:
            print(f"  WARNING: No Metrics sheet in {bn}")
            wb.close()
            continue

        ws = wb[metrics_name]

        # ── Human section: rows 3-10 (header at row 2, data at 3-9, overall at 10) ──
        if example not in human_examples_seen:
            human_section = read_section(ws, start_row=3, end_row=11)
            accumulate(aggregated[HUMAN_GRADER_LABEL], human_section)
            human_examples_seen.add(example)
            human_flag = "  [+human]"
        else:
            human_flag = ""

        # ── LLM section: find header row at or after row 12, then read data ──
        header_row = None
        for rn in range(12, 30):
            cell_val = ws.cell(rn, 1).value
            if cell_val and str(cell_val).strip().lower() == "type":
                header_row = rn
                break

        if header_row is None:
            print(f"  WARNING: No LLM header row in {bn}")
            wb.close()
            continue

        llm_section = read_section(
            ws, start_row=header_row + 1, end_row=header_row + 20
        )
        accumulate(aggregated[llm], llm_section)

        wb.close()
        print(
            f"  {example:35} | {llm:25} | "
            f"{len(llm_section)} LLM types, {len(human_section if human_flag else {})} human{human_flag}"
        )

    # ── Build output rows ─────────────────────────────────────────────────────
    GRADER_ORDER = [
        HUMAN_GRADER_LABEL,
        "Claude4.5Sonnet",
        "GPT-5.5",
        "Gemini3.1ProPreview",
    ]
    output_rows = []

    for grader in GRADER_ORDER:
        if grader not in aggregated:
            print(f"  WARNING: No data for grader '{grader}'")
            continue

        total_n = total_tp = total_fp = total_fn = 0.0

        for etype in TARGET_TYPES:
            d = aggregated[grader][etype]
            tp, fp, fn, n = d["TP"], d["FP"], d["FN"], d["N"]
            precision, recall, f1 = compute_metrics(tp, fp, fn)
            output_rows.append(
                {
                    "Grader": grader,
                    "ElementType": etype,
                    "TotalN": round(n, 4),
                    "TotalTP": round(tp, 4),
                    "TotalFP": round(fp, 4),
                    "TotalFN": round(fn, 4),
                    "TotalElements": round(n + fp, 4),
                    "Precision": fmt(precision),
                    "Recall": fmt(recall),
                    "F1": fmt(f1),
                }
            )
            total_n += n
            total_tp += tp
            total_fp += fp
            total_fn += fn

        # Overall row for this grader
        precision, recall, f1 = compute_metrics(total_tp, total_fp, total_fn)
        output_rows.append(
            {
                "Grader": grader,
                "ElementType": "Overall",
                "TotalN": round(total_n, 4),
                "TotalTP": round(total_tp, 4),
                "TotalFP": round(total_fp, 4),
                "TotalFN": round(total_fn, 4),
                "TotalElements": round(total_n + total_fp, 4),
                "Precision": fmt(precision),
                "Recall": fmt(recall),
                "F1": fmt(f1),
            }
        )

    # ── Write CSV ─────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    fields = [
        "Grader",
        "ElementType",
        "TotalN",
        "TotalTP",
        "TotalFP",
        "TotalFN",
        "TotalElements",
        "Precision",
        "Recall",
        "F1",
    ]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(output_rows)

    # ── Console summary ───────────────────────────────────────────────────────
    print("\n" + "=" * 100)
    print(
        f"{'Grader':25} {'ElementType':20} {'N':>7} {'TP':>7} {'FP':>7} {'FN':>7}  {'Prec':>8} {'Rec':>8} {'F1':>8}"
    )
    print("=" * 100)
    prev_grader = None
    for r in output_rows:
        if r["Grader"] != prev_grader:
            if prev_grader is not None:
                print()
            prev_grader = r["Grader"]
        marker = "  <-- Overall" if r["ElementType"] == "Overall" else ""
        print(
            f"{r['Grader']:25} {r['ElementType']:20} "
            f"{r['TotalN']:>7} {r['TotalTP']:>7} {r['TotalFP']:>7} {r['TotalFN']:>7}  "
            f"{str(r['Precision']):>8} {str(r['Recall']):>8} {str(r['F1']):>8}{marker}"
        )

    print(f"\nOutput written to: {out_path}")


if __name__ == "__main__":
    main()
