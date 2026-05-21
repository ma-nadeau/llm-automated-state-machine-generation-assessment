#!/usr/bin/env python3
"""
Compute the distribution_of_state_machine_elements_across_grading_sheets.csv.

Reads the reference sheet (index 0) of each CombinedHumanGradingVsClaude4.5Sonnet
workbook (one per example) and counts ground-truth elements by type, excluding
rows labelled 'additional elements' in column B (FP template rows).

Output: _Figures/data/distribution_of_state_machine_elements_across_grading_sheets.csv
"""

import csv
import glob
import os
import re
from collections import defaultdict

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
OUTPUT_DIR = os.path.join(BASE_DIR, "_Figures", "data")
OUTPUT_FILE = "distribution_of_state_machine_elements_across_grading_sheets.csv"

ELEMENT_TYPES = [
    "States",
    "Transitions",
    "Guards",
    "Actions",
    "Hierarchical States",
    "Parallel Regions",
    "History States",
]

TYPE_ALIASES = {
    "state": "States",
    "states": "States",
    "transition": "Transitions",
    "transitions": "Transitions",
    "guard": "Guards",
    "composite state": "Hierarchical States",
    "composite states": "Hierarchical States",
    "action": "Actions",
    "region": "Parallel Regions",
    "history state": "History States",
    "history states": "History States",
}

EXAMPLE_LABELS = {
    "Automatic Bread Maker": "Bread Maker",
    "Digital Chess Clock": "Chess Clock",
    "Dishwasher": "Dishwasher",
    "Printer": "Printer",
    "SSC7": "SSC7",
    "Spa Manager": "Spa Manager",
    "Train Automation System": "Train Automation",
    "Thermomix TM6": "Thermomix TM6",
    "Wumple": "W-UMPLE",
}


def normalize_type(raw):
    if raw is None:
        return None
    return TYPE_ALIASES.get(str(raw).strip().lower())


def read_reference_counts(ws):
    counts = defaultdict(int)
    for rn in range(2, ws.max_row + 1):
        col_a = ws.cell(rn, 1).value
        col_b = ws.cell(rn, 2).value
        if col_a is None and col_b is None:
            break
        if col_b is not None and str(col_b).strip().lower() == "additional elements":
            continue
        norm = normalize_type(col_a)
        if norm:
            counts[norm] += 1
    return dict(counts)


def main():
    pattern = os.path.join(
        BASE_DIR,
        "*/Grading/2 stage/6-examples/"
        "*_CombinedHumanGradingVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration.xlsx",
    )
    files = sorted(
        f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~")
    )
    if not files:
        print("ERROR: No matching files found.")
        return

    # example_name -> {type -> count}
    data = {}
    for fp in files:
        raw_example = os.path.basename(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(fp))))
        )
        label = EXAMPLE_LABELS.get(raw_example, raw_example)
        wb = openpyxl.load_workbook(fp, data_only=True)
        counts = read_reference_counts(wb[wb.sheetnames[0]])
        wb.close()
        data[label] = counts
        print(f"  {label}: {dict(counts)}")

    # Ordered list of system labels
    systems = [EXAMPLE_LABELS[k] for k in EXAMPLE_LABELS if EXAMPLE_LABELS[k] in data]

    # Build rows
    rows = []
    row_totals = {}
    for etype in ELEMENT_TYPES:
        row = [etype] + [data[s].get(etype, 0) for s in systems]
        row_total = sum(row[1:])
        row.append(row_total)
        row_totals[etype] = row_total
        rows.append(row)

    col_totals = [sum(data[s].get(e, 0) for e in ELEMENT_TYPES) for s in systems]
    grand_total = sum(col_totals)
    total_row = ["Total"] + col_totals + [grand_total]

    # Write CSV
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    header = ["Component"] + systems + ["Total"]
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
        w.writerow(total_row)

    # Console summary
    col_w = 16
    print("\n" + "-" * (12 + col_w * (len(systems) + 1)))
    print(f"{'Component':<20}", end="")
    for s in systems:
        print(f"{s:>{col_w}}", end="")
    print(f"{'Total':>{col_w}}")
    print("-" * (20 + col_w * (len(systems) + 1)))
    for row in rows:
        print(f"{row[0]:<20}", end="")
        for v in row[1:]:
            print(f"{v:>{col_w}}", end="")
        print()
    print("-" * (20 + col_w * (len(systems) + 1)))
    print(f"{'Total':<20}", end="")
    for v in total_row[1:]:
        print(f"{v:>{col_w}}", end="")
    print()
    print(f"\nOutput written to: {out_path}")


if __name__ == "__main__":
    main()
