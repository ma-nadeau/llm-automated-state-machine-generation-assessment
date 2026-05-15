"""
extract_cohens_kappa.py

Discovers all Excel workbooks that represent combined human-vs-LLM grading
comparisons, reads the weighted Cohen's Kappa values from the 6th sheet
("Weighted Cohens Kappa"), and writes:

  aggregated_cohens_kappa_raw.csv – one row per (file, category)

Usage:
    python extract_cohens_kappa.py [--root <evaluations_dir>] [--out <output_dir>]
"""

import argparse
import logging
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FILE_PATTERN = "*CombinedHumanGrading*AutoGrading*.xlsx"

# Regex to pull the LLM grader name out of the filename.
# Matches: CombinedHumanGradingVs<LLM>AutoGrading
LLM_RE = re.compile(r"CombinedHumanGradingVs(.+?)AutoGrading", re.IGNORECASE)

KAPPA_SHEET_INDEX = 5          # 0-based; sheet 6
COL_V = 21                     # 0-based column index for metric label
COL_W = 22                     # 0-based column index for metric value

KAPPA_ROW_LABEL = "Cohen's Kappa"
METRICS_PREFIX = "Metrics "
# Older 1-stage sheets use the generic header "Metric" with no category name.
# The blocks appear in CATEGORY_ORDER sequence, so we assign categories by position.
GENERIC_METRIC_HEADER = "Metric"

# Canonical category ordering used to name positional blocks in older 1-stage sheets
CATEGORY_ORDER = [
    "Global",
    "State",
    "Transition",
    "Composite State",
    "Guard",
    "Action",
    "History State",
    "Region",
]

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s  %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_llm(filename: str) -> str | None:
    m = LLM_RE.search(filename)
    return m.group(1) if m else None


def _extract_project(filename: str) -> str:
    # Everything before the first "_Grading_" token
    return filename.split("_Grading_")[0]


def _as_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if s in ("#DIV/0!", "#N/A", "#VALUE!", "#REF!", "#NUM!", "#NAME?", ""):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Core extraction
# ---------------------------------------------------------------------------

def extract_kappa_from_workbook(path: Path) -> list[dict]:
    """
    Returns a list of dicts, one per category found in the kappa sheet.
    Returns an empty list and logs a warning on any recoverable error.
    """
    filename = path.name
    llm_grader = _extract_llm(filename)
    project = _extract_project(filename)

    if llm_grader is None:
        log.warning("SKIP  Could not parse LLM grader from filename: %s", filename)
        return []

    # ---- open workbook -------------------------------------------------
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        log.error("SKIP  Failed to open workbook %s: %s", path, exc)
        return []

    # ---- locate sheet --------------------------------------------------
    if len(wb.sheetnames) <= KAPPA_SHEET_INDEX:
        log.warning(
            "SKIP  Workbook has only %d sheet(s), expected at least %d: %s",
            len(wb.sheetnames),
            KAPPA_SHEET_INDEX + 1,
            filename,
        )
        return []

    ws = wb.worksheets[KAPPA_SHEET_INDEX]
    sheet_name = ws.title

    if "kappa" not in sheet_name.lower():
        log.warning(
            "WARN  Sheet 6 is named %r (expected a 'kappa' sheet) in %s",
            sheet_name,
            filename,
        )

    # ---- scan rows for (Metrics <Category> … Cohen's Kappa <value>) ----
    # Structure:
    #   Col V = "Metrics Global"  → starts a new category block
    #   Col V = "Cohen's Kappa"   → the value for the current category is in Col W
    rows = list(ws.iter_rows(values_only=True))

    results: list[dict] = []
    current_category: str | None = None
    positional_block_index: int = 0   # for sheets that use generic "Metric" headers
    uses_generic_headers: bool = False

    for row_idx, row in enumerate(rows, start=1):
        padded = list(row) + [None] * (COL_W + 2)
        v_val = padded[COL_V]
        w_val = padded[COL_W]

        if not isinstance(v_val, str):
            continue

        v_stripped = v_val.strip()

        if v_stripped.startswith(METRICS_PREFIX):
            current_category = v_stripped[len(METRICS_PREFIX):]
            continue

        if v_stripped == GENERIC_METRIC_HEADER:
            # Older sheet format — assign category by block position
            uses_generic_headers = True
            if positional_block_index < len(CATEGORY_ORDER):
                current_category = CATEGORY_ORDER[positional_block_index]
            else:
                current_category = f"Unknown_{positional_block_index}"
                log.warning(
                    "WARN  More metric blocks than known categories at row %d in %s",
                    row_idx,
                    filename,
                )
            positional_block_index += 1
            continue

        if v_stripped == KAPPA_ROW_LABEL:
            if current_category is None:
                log.warning(
                    "WARN  Found '%s' row at row %d before any metric header in %s",
                    KAPPA_ROW_LABEL,
                    row_idx,
                    filename,
                )
                continue

            kappa = _as_float(w_val)
            is_div_zero = isinstance(w_val, str) and "#DIV/0!" in w_val

            if kappa is None and not is_div_zero and w_val is not None:
                log.warning(
                    "WARN  Unexpected kappa value %r for category %r at row %d in %s",
                    w_val,
                    current_category,
                    row_idx,
                    filename,
                )

            note = ""
            if is_div_zero:
                note = "undefined (division by zero)"
            elif uses_generic_headers:
                note = "category inferred by block position (generic header sheet)"

            results.append(
                {
                    "project": project,
                    "file": filename,
                    "llm_grader": llm_grader,
                    "category": current_category,
                    "weighted_cohens_kappa": kappa,
                    "raw_value": str(w_val) if w_val is not None else "",
                    "note": note,
                }
            )
            current_category = None  # reset; next header starts a new block

    if not results:
        log.warning("WARN  No kappa values extracted from %s", filename)

    return results


# ---------------------------------------------------------------------------
# Validation: column W consistency check
# ---------------------------------------------------------------------------

def _validate_col_w_consistency(all_rows: list[dict]) -> None:
    """
    The sheet does not embed an LLM identifier in column W; the LLM name comes
    from the filename. This function checks that every file parsed to a single
    consistent llm_grader value (trivially true for well-formed filenames) and
    flags any file that produced duplicate category entries.
    """
    by_file: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        by_file[row["file"]].append(row)

    for fname, rows in by_file.items():
        llms = {r["llm_grader"] for r in rows}
        if len(llms) > 1:
            log.error(
                "INCONSISTENCY  File %s resolved to multiple LLM grader values: %s",
                fname,
                llms,
            )
        cats = [r["category"] for r in rows]
        duplicates = {c for c in cats if cats.count(c) > 1}
        if duplicates:
            log.warning(
                "WARN  File %s has duplicate category entries: %s",
                fname,
                duplicates,
            )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--root",
        default=str(Path(__file__).parent.parent),
        help="Root Evaluations directory to search (default: parent of this script)",
    )
    parser.add_argument(
        "--out",
        default=str(Path(__file__).parent.parent / "_Figures" / "PaperExperimentData"),
        help="Output directory for CSV files",
    )
    args = parser.parse_args()

    root = Path(args.root)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    # ---- discover files ------------------------------------------------
    all_files = sorted(root.rglob(FILE_PATTERN))
    log.info("Found %d candidate file(s) matching pattern.", len(all_files))

    if not all_files:
        log.error("No files found. Check --root path: %s", root)
        sys.exit(1)

    # ---- extract -------------------------------------------------------
    all_rows: list[dict] = []
    processed, skipped = 0, 0

    for path in all_files:
        rows = extract_kappa_from_workbook(path)
        if rows:
            all_rows.extend(rows)
            processed += 1
        else:
            skipped += 1

    if not all_rows:
        log.error("No data extracted from any file. Aborting.")
        sys.exit(1)

    _validate_col_w_consistency(all_rows)

    # ---- raw CSV -------------------------------------------------------
    raw_df = pd.DataFrame(all_rows)
    raw_path = out_dir / "aggregated_cohens_kappa_raw.csv"
    raw_df.to_csv(raw_path, index=False)
    log.info("Wrote raw CSV  → %s  (%d rows)", raw_path, len(raw_df))

    # ---- console summary ----------------------------------------------
    detected_llms = sorted(raw_df["llm_grader"].unique())
    detected_cats = sorted(raw_df["category"].unique())
    undefined_count = raw_df["weighted_cohens_kappa"].isna().sum()

    print("\n" + "=" * 60)
    print("EXTRACTION COMPLETE")
    print("=" * 60)
    print(f"  Files processed : {processed}")
    print(f"  Files skipped   : {skipped}")
    print(f"  Total rows      : {len(raw_df)}")
    print(f"  Undefined kappa : {undefined_count}  (logged as NaN; #DIV/0! in sheet)")
    print(f"  Detected LLMs   : {', '.join(detected_llms)}")
    print(f"  Detected cats   : {', '.join(detected_cats)}")
    print(f"\n  Raw CSV → {raw_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
