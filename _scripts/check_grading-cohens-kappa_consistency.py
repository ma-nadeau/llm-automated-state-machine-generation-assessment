"""
check_grading_consistency.py

For every Excel workbook whose filename:
  - does NOT contain "- Template"
  - DOES contain "_CombinedHumanGrading" OR "_HumanGrading_"

Compare:
  Sheet 2 Col C  vs  Sheet 6 Col C   (human / rater-1 grades)
  Sheet 3 Col C  vs  Sheet 6 Col D   (LLM / rater-2 grades)

Sheet 6 is expected to be "Weighted Cohens Kappa" (or an alias).
Any mismatch is written to grading_consistency_report.csv.
"""

import csv
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path("/Users/marc-antoinenadeau/Desktop/B.Eng Software Engineering /7th Semester/Capstone/Code")
OUTPUT_CSV = ROOT / "Evaluations/_Figures/PaperExperimentData/grading-cohens-kappa_consistency_report.csv"

KAPPA_ALIASES = {"weighted cohens kappa", "weighted cohen's kappa", "cohens kappa", "cohen's kappa", "kappa"}

CSV_HEADERS = [
    "file_name",
    "sheet_2_name",
    "sheet_3_name",
    "sheet_6_name",
    "comparison",
    "row_number",
    "source_value",
    "kappa_sheet_value",
    "discrepancy",
]


def find_target_files(root: Path) -> list[Path]:
    files = []
    for path in root.rglob("*.xlsx"):
        name = path.name
        if name.startswith("~$"):
            continue
        if "- Template" in name:
            continue
        if "_CombinedHumanGrading" in name or "_HumanGrading_" in name:
            files.append(path)
    return sorted(files)


def read_column(ws, col_index: int) -> list:
    """Return all cell values in col_index (1-based) as a list."""
    return [row[0] for row in ws.iter_rows(min_col=col_index, max_col=col_index, values_only=True)]


def effective_length(values: list) -> int:
    """Index of last non-None value + 1 (i.e. how many rows actually have data)."""
    for i in range(len(values) - 1, -1, -1):
        if values[i] is not None:
            return i + 1
    return 0


def normalize(v):
    """Normalize a cell value for comparison (strip whitespace from strings)."""
    if isinstance(v, str):
        return v.strip()
    return v


def compare_columns(
    file_name: str,
    sheet2_name: str,
    sheet3_name: str,
    sheet6_name: str,
    s2_col_c: list,
    s3_col_c: list,
    s6_col_c: list,
    s6_col_d: list,
) -> list[dict]:
    discrepancies = []

    # Determine effective data length (skip pure-None trailing rows)
    # Both pairs must agree on the same row universe; use the longer effective length
    # so we catch "missing" rows too.
    len_s2_s6c = max(effective_length(s2_col_c), effective_length(s6_col_c))
    len_s3_s6d = max(effective_length(s3_col_c), effective_length(s6_col_d))

    # Comparison 1: Sheet 2 Col C  vs  Sheet 6 Col C
    for row_idx in range(1, len_s2_s6c):          # skip header (index 0 = row 1)
        v_src  = normalize(s2_col_c[row_idx] if row_idx < len(s2_col_c) else None)
        v_kapp = normalize(s6_col_c[row_idx] if row_idx < len(s6_col_c) else None)
        if v_src != v_kapp:
            discrepancies.append({
                "file_name":        file_name,
                "sheet_2_name":     sheet2_name,
                "sheet_3_name":     sheet3_name,
                "sheet_6_name":     sheet6_name,
                "comparison":       f"{sheet2_name} Col C  vs  {sheet6_name} Col C",
                "row_number":       row_idx + 1,     # 1-based Excel row
                "source_value":     v_src,
                "kappa_sheet_value": v_kapp,
                "discrepancy":      f"Sheet 2 Col C = {v_src!r}, Sheet 6 Col C = {v_kapp!r}",
            })

    # Comparison 2: Sheet 3 Col C  vs  Sheet 6 Col D
    for row_idx in range(1, len_s3_s6d):
        v_src  = normalize(s3_col_c[row_idx] if row_idx < len(s3_col_c) else None)
        v_kapp = normalize(s6_col_d[row_idx] if row_idx < len(s6_col_d) else None)
        if v_src != v_kapp:
            discrepancies.append({
                "file_name":        file_name,
                "sheet_2_name":     sheet2_name,
                "sheet_3_name":     sheet3_name,
                "sheet_6_name":     sheet6_name,
                "comparison":       f"{sheet3_name} Col C  vs  {sheet6_name} Col D",
                "row_number":       row_idx + 1,
                "source_value":     v_src,
                "kappa_sheet_value": v_kapp,
                "discrepancy":      f"Sheet 3 Col C = {v_src!r}, Sheet 6 Col D = {v_kapp!r}",
            })

    return discrepancies


def process_file(path: Path) -> tuple[list[dict], list[str]]:
    """Return (discrepancies, warnings)."""
    warnings = []
    discrepancies = []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        warnings.append(f"  SKIP — could not open: {e}")
        return discrepancies, warnings

    sheets = wb.sheetnames
    if len(sheets) < 6:
        warnings.append(f"  SKIP — only {len(sheets)} sheet(s); need at least 6")
        wb.close()
        return discrepancies, warnings

    sheet6_name = sheets[5]
    if sheet6_name.lower() not in KAPPA_ALIASES:
        warnings.append(
            f"  WARNING — sheet 6 is '{sheet6_name}', not a recognised Kappa alias; proceeding anyway"
        )

    ws2 = wb.worksheets[1]
    ws3 = wb.worksheets[2]
    ws6 = wb.worksheets[5]

    s2_col_c = read_column(ws2, 3)
    s3_col_c = read_column(ws3, 3)
    s6_col_c = read_column(ws6, 3)
    s6_col_d = read_column(ws6, 4)

    wb.close()

    discrepancies = compare_columns(
        file_name=path.name,
        sheet2_name=sheets[1],
        sheet3_name=sheets[2],
        sheet6_name=sheet6_name,
        s2_col_c=s2_col_c,
        s3_col_c=s3_col_c,
        s6_col_c=s6_col_c,
        s6_col_d=s6_col_d,
    )
    return discrepancies, warnings


def main():
    files = find_target_files(ROOT)
    if not files:
        print("No target files found.")
        return

    print(f"Found {len(files)} target file(s).\n")

    all_discrepancies = []
    for path in files:
        print(f"Processing: {path.name}")
        disc, warns = process_file(path)
        for w in warns:
            print(w)
        if disc:
            print(f"  {len(disc)} discrepancy(ies) found")
        else:
            print("  OK — no discrepancies")
        all_discrepancies.extend(disc)

    print(f"\nTotal discrepancies: {len(all_discrepancies)}")
    print(f"Writing report to: {OUTPUT_CSV}")

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(all_discrepancies)

    print("Done.")


if __name__ == "__main__":
    main()
