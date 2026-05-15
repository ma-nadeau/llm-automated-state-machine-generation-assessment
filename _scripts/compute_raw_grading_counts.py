#!/usr/bin/env python3
"""
Compute grading counts (N, TP, FP, FN) directly from the raw data sheets,
bypassing the Metrics sheet (which contains formula-range errors in some workbooks).

Two root-cause issues found in the source files:
  1. Digital Chess Clock  – LLM-section N formulas still reference the old range
     $A$2:$A$44 after two ground-truth rows (State/Transition) were appended at
     rows 45-46. Fixed by reading N from the reference sheet instead.
  2. SSC7                 – GPT-5.5 / Gemini grading sheets mislabel row 53
     (Action "error()") as a Transition. The reference sheet is authoritative.
  3. Train Automation System – Claude LLM grading sheet is missing 5 Action rows
     that exist in the reference; those rows implicitly contribute TP = 0.

Fix: N is always taken from the reference sheet (workbook sheet index 0).
     TP and FP are summed from the grading sheets (indices 1 and 2).

Workbook layout (consistent across all 27 files):
  Sheet index 0  – reference / ground-truth list  (col A: type, col B: element name)
  Sheet index 1  – Human grading sheet            (same across all 3 LLM files)
  Sheet index 2  – LLM auto-grading sheet
  Sheet index 3  – Metrics  (NOT read by this script)

Row structure inside each sheet:
  Row 1        – header
  Rows 2..X    – ground-truth items  (col B ≠ "additional elements")
  Rows X+1..Y  – FP template rows   (col B  = "additional elements")

Outputs (written to _Figures/PaperExperimentData/):
  2stage_6examples_PerExample_RawCounts_CombinedHumanVsLLM.csv
      Example, Grader, ElementType, TotalN, TotalTP, TotalFP, TotalFN, TotalElements, Precision, Recall, F1
      TotalElements = TotalN + TotalFP (all expected + all spurious generated elements)

  2stage_6examples_Metrics_CombinedHumanVsLLM.csv  [overwrites previous version]
      Grader, ElementType, TotalN, TotalTP, TotalFP, TotalFN, TotalElements, Precision, Recall, F1
      (aggregated across all 9 examples)
"""

import csv
import glob
import os
import re
from collections import defaultdict

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR    = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR  = os.path.join(BASE_DIR, "_Figures", "PaperExperimentData")

TARGET_TYPES = [
    "Action",
    "Composite State",
    "Guard",
    "History State",
    "Region",
    "State",
    "Transition",
]

TYPE_ALIASES = {
    "action":           "Action",
    "composite state":  "Composite State",
    "composite states": "Composite State",
    "guard":            "Guard",
    "history state":    "History State",
    "history states":   "History State",
    "region":           "Region",
    "state":            "State",
    "states":           "State",
    "transition":       "Transition",
    "transitions":      "Transition",
}

HUMAN_GRADER_LABEL = "Human"
GRADER_ORDER = [HUMAN_GRADER_LABEL, "Claude4.5Sonnet", "GPT-5.5", "Gemini3.1ProPreview"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_type(raw):
    if raw is None:
        return None
    return TYPE_ALIASES.get(str(raw).strip().lower())


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def is_fp_row(col_b):
    return col_b is not None and str(col_b).strip().lower() == "additional elements"


def compute_metrics(tp, fp, fn):
    precision = tp / (tp + fp) if (tp + fp) > 0 else None
    recall    = tp / (tp + fn) if (tp + fn) > 0 else None
    if precision is not None and recall is not None and (precision + recall) > 0:
        f1 = 2 * precision * recall / (precision + recall)
    else:
        f1 = None
    return precision, recall, f1


def fmt(val, decimals=6):
    return round(val, decimals) if val is not None else "N/A"


# ── Sheet readers ─────────────────────────────────────────────────────────────

def read_reference_n(ws):
    """
    Count ground-truth N per element type from the reference sheet (index 0).
    Rows with col B == 'additional elements' are skipped (FP template section).
    Returns dict: norm_type -> int
    """
    counts = defaultdict(int)
    for rn in range(2, ws.max_row + 1):
        col_a = ws.cell(rn, 1).value
        col_b = ws.cell(rn, 2).value
        if col_a is None and col_b is None:
            break
        if is_fp_row(col_b):
            continue
        norm = normalize_type(col_a)
        if norm is not None:
            counts[norm] += 1
    return dict(counts)


def read_grading_tp_fp(ws):
    """
    Sum TP and FP scores from a grading sheet (index 1 or 2).
    Ground-truth rows (col B ≠ 'additional elements') → TP.
    FP rows (col B == 'additional elements')           → FP.
    Element type is taken from the grading sheet's col A.
    Returns (tp_dict, fp_dict): norm_type -> float
    """
    tp = defaultdict(float)
    fp = defaultdict(float)
    for rn in range(2, ws.max_row + 1):
        col_a = ws.cell(rn, 1).value
        col_b = ws.cell(rn, 2).value
        col_c = ws.cell(rn, 3).value
        if col_a is None and col_b is None:
            break
        if col_a is None:
            continue
        norm = normalize_type(col_a)
        if norm is None:
            continue
        score = safe_float(col_c) or 0.0
        if is_fp_row(col_b):
            fp[norm] += score
        else:
            tp[norm] += score
    return dict(tp), dict(fp)


def combine(ref_n, tp_dict, fp_dict):
    """
    Build per-type result using reference N and grading-sheet TP/FP.
    FN = N - TP.
    """
    result = {}
    for etype in TARGET_TYPES:
        n  = float(ref_n.get(etype, 0))
        tp = tp_dict.get(etype, 0.0)
        fp = fp_dict.get(etype, 0.0)
        result[etype] = {"N": n, "TP": tp, "FP": fp, "FN": n - tp}
    return result


def build_overall(type_data):
    total = {"N": 0.0, "TP": 0.0, "FP": 0.0, "FN": 0.0}
    for d in type_data.values():
        for k in total:
            total[k] += d[k]
    return total


def make_row(grader, etype, d):
    tp, fp, fn, n = d["TP"], d["FP"], d["FN"], d["N"]
    p, r, f1 = compute_metrics(tp, fp, fn)
    return {
        "Grader":         grader,
        "ElementType":    etype,
        "TotalN":         round(n,  4),
        "TotalTP":        round(tp, 4),
        "TotalFP":        round(fp, 4),
        "TotalFN":        round(fn, 4),
        "TotalElements":  round(n + fp, 4),
        "Precision":      fmt(p),
        "Recall":         fmt(r),
        "F1":             fmt(f1),
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    pattern = os.path.join(
        BASE_DIR,
        "*/Grading/2 stage/6-examples/"
        "*_CombinedHumanGradingVs*AutoGrading_Claude4.5SonnetGeneration.xlsx",
    )
    files = sorted(
        f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~")
    )
    if not files:
        print("ERROR: No matching files found.")
        return
    print(f"Found {len(files)} files.\n")

    # per_example[example][grader][etype] = {N, TP, FP, FN}
    per_example = defaultdict(lambda: defaultdict(dict))
    human_seen  = set()
    warnings    = []

    for filepath in files:
        bn  = os.path.basename(filepath)
        m   = re.search(r"CombinedHumanGradingVs(.+?)AutoGrading", bn)
        if not m:
            print(f"  WARNING: Cannot parse LLM name from {bn}")
            continue
        llm     = m.group(1)
        example = os.path.basename(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(filepath))))
        )

        wb     = openpyxl.load_workbook(filepath, data_only=True)
        sheets = wb.sheetnames
        if len(sheets) < 3:
            print(f"  WARNING: Too few sheets in {bn}")
            wb.close()
            continue

        # Reference N (same reference for both human and LLM graders)
        ref_n = read_reference_n(wb[sheets[0]])

        # LLM grading (sheet index 2)
        llm_tp, llm_fp = read_grading_tp_fp(wb[sheets[2]])
        per_example[example][llm] = combine(ref_n, llm_tp, llm_fp)

        # Warn if any type has more grading-sheet rows than reference (suggests
        # a labeling error: extra rows of that type in the grading sheet)
        for etype in TARGET_TYPES:
            ref_count  = ref_n.get(etype, 0)
            grade_rows = sum(
                1 for rn in range(2, wb[sheets[2]].max_row + 1)
                if normalize_type(wb[sheets[2]].cell(rn, 1).value) == etype
                and not is_fp_row(wb[sheets[2]].cell(rn, 2).value)
                and wb[sheets[2]].cell(rn, 1).value is not None
            )
            if grade_rows != ref_count:
                warnings.append(
                    f"  N mismatch [{example} / {llm} / {etype}]: "
                    f"reference={ref_count}, grading-sheet={grade_rows}"
                )

        # Human grading (sheet index 1) — read once per example
        if example not in human_seen:
            human_tp, human_fp = read_grading_tp_fp(wb[sheets[1]])
            per_example[example][HUMAN_GRADER_LABEL] = combine(ref_n, human_tp, human_fp)
            human_seen.add(example)
            human_tag = "  [+human]"
        else:
            human_tag = ""

        wb.close()
        print(f"  {example:35} | {llm:25}{human_tag}")

    if warnings:
        print("\n--- GRADING-SHEET N DISCREPANCIES (ref N ≠ rows in grading sheet) ---")
        for w in warnings:
            print(w)

    # ── Build per-example rows ────────────────────────────────────────────────
    per_example_rows = []
    for example in sorted(per_example):
        for grader in GRADER_ORDER:
            if grader not in per_example[example]:
                continue
            type_data = per_example[example][grader]
            for etype in TARGET_TYPES:
                row = make_row(grader, etype, type_data[etype])
                row["Example"] = example
                per_example_rows.append(row)
            row = make_row(grader, "Overall", build_overall(type_data))
            row["Example"] = example
            per_example_rows.append(row)

    # ── Build aggregated rows ─────────────────────────────────────────────────
    agg = defaultdict(lambda: defaultdict(lambda: {"N":0.,"TP":0.,"FP":0.,"FN":0.}))
    for example in per_example:
        for grader in per_example[example]:
            for etype in TARGET_TYPES:
                d = per_example[example][grader][etype]
                for k in ("N", "TP", "FP", "FN"):
                    agg[grader][etype][k] += d[k]

    aggregated_rows = []
    for grader in GRADER_ORDER:
        if grader not in agg:
            continue
        for etype in TARGET_TYPES:
            aggregated_rows.append(make_row(grader, etype, agg[grader][etype]))
        aggregated_rows.append(make_row(grader, "Overall", build_overall(agg[grader])))

    # ── Write CSVs ────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    per_ex_path = os.path.join(
        OUTPUT_DIR, "2stage_6examples_PerExample_RawCounts_CombinedHumanVsLLM.csv"
    )
    fields_ex = [
        "Example", "Grader", "ElementType",
        "TotalN", "TotalTP", "TotalFP", "TotalFN", "TotalElements", "Precision", "Recall", "F1",
    ]
    with open(per_ex_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields_ex)
        w.writeheader()
        w.writerows(per_example_rows)

    agg_path = os.path.join(
        OUTPUT_DIR, "2stage_6examples_Metrics_CombinedHumanVsLLM.csv"
    )
    fields_agg = [
        "Grader", "ElementType",
        "TotalN", "TotalTP", "TotalFP", "TotalFN", "TotalElements", "Precision", "Recall", "F1",
    ]
    with open(agg_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields_agg)
        w.writeheader()
        w.writerows(aggregated_rows)

    # ── Console summary ───────────────────────────────────────────────────────
    print("\n" + "=" * 105)
    print("PER-EXAMPLE COUNTS (N from reference sheet; TP/FP from grading sheets)")
    print("=" * 105)
    print(f"  {'Example':30} {'Grader':22} {'Type':20} {'N':>5} {'TP':>6} {'FP':>5} {'FN':>6}  {'P':>7} {'R':>7} {'F1':>7}")
    print("  " + "-" * 103)
    prev_ex = None
    for r in per_example_rows:
        if r["Example"] != prev_ex:
            if prev_ex is not None:
                print()
            prev_ex = r["Example"]
        marker = " *" if r["ElementType"] == "Overall" else "  "
        print(
            f"  {r['Example']:30} {r['Grader']:22} {r['ElementType']:20} "
            f"{r['TotalN']:>5} {r['TotalTP']:>6} {r['TotalFP']:>5} {r['TotalFN']:>6}  "
            f"{str(r['Precision']):>7} {str(r['Recall']):>7} {str(r['F1']):>7}{marker}"
        )

    print("\n" + "=" * 100)
    print("AGGREGATED ACROSS ALL 9 EXAMPLES")
    print("=" * 100)
    print(f"{'Grader':25} {'ElementType':20} {'N':>7} {'TP':>7} {'FP':>6} {'FN':>7}  {'Prec':>8} {'Rec':>8} {'F1':>8}")
    print("-" * 100)
    prev_grader = None
    for r in aggregated_rows:
        if r["Grader"] != prev_grader:
            if prev_grader is not None:
                print()
            prev_grader = r["Grader"]
        marker = "  <-- Overall" if r["ElementType"] == "Overall" else ""
        print(
            f"{r['Grader']:25} {r['ElementType']:20} "
            f"{r['TotalN']:>7} {r['TotalTP']:>7} {r['TotalFP']:>6} {r['TotalFN']:>7}  "
            f"{str(r['Precision']):>8} {str(r['Recall']):>8} {str(r['F1']):>8}{marker}"
        )

    print(f"\nOutputs written to:\n  {per_ex_path}\n  {agg_path}")


if __name__ == "__main__":
    main()
