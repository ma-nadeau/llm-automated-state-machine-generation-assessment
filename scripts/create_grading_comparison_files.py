#!/usr/bin/env python3
"""
Create grading-comparison Excel templates from existing CombinedHumanGradingVsClaude4.5SonnetGrading_Claude4.5SonnetGeneration.xlsx.

Experiment type: keep the generated state machines FIXED (Claude 4.5 Sonnet generation)
and vary the AUTO-GRADER model (e.g. GPT-5.5, Gemini 3.1 Pro Preview), comparing each new grader
against the Claude 4.5 Sonnet baseline grader.

Output naming convention:
  {Example}_Grading_2-stage_{date}_{N}-examples_
    {NewGrader}AutoGradingVSClaude4.5SonnetAutoGrading-Claude4.5SonnetGeneration.xlsx

Changes made inside each new workbook:
  • Ground-truth sheet        – unchanged
  • Human-grading sheet       – renamed to "{New Grader} Grading";
                                grade column (C) cleared (placeholder for new grader data);
                                notes column (D) cleared;
                                column C header updated
  • LLM-grading sheet         – renamed to "Claude 4.5 Sonnet Grading";
                                data preserved (Claude 4.5 Sonnet baseline grader)
                                column C header updated
  • Metrics                   – A1 = "Graded by {NewGrader}"
                                A12 = "Graded by Claude 4.5 Sonnet"
                                column J: J4='Generation LLM', J5='Claude 4.5 Sonnet',
                                          J7='Grading LLMs', J8={NewGrader}, J9='Claude 4.5 Sonnet'
  • Inter-rater               – column C/D headers updated
  • Weighted Cohens Kappa     – column C/D headers updated
  • All formula strings       – old sheet-name references replaced with new names

Output path: {date_dir}/{filename}.xlsx  (same level as the other xlsx files)

Usage:
  # Default: GPT-5.5 and Gemini 3.1 Pro Preview grading-comparison files for all 9 examples
  python create_grading_comparison_files.py

  # Specific grader(s) only
  python create_grading_comparison_files.py --graders "GPT-5.5"
  python create_grading_comparison_files.py --graders "GPT-5.5" "Gemini 3.1 Pro Preview"

  # Override the embedded date
  python create_grading_comparison_files.py --date 2026-05-04
"""

from __future__ import annotations

import argparse
import glob
import os
import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent

# The generation model that is FIXED across all grading-comparison experiments
GENERATION_LLM_DISPLAY = "Claude 4.5 Sonnet"  # human-readable
GENERATION_LLM_FILE = "Claude4.5Sonnet"  # used in file/sheet names (no spaces)

# Default graders to create comparison files for
DEFAULT_GRADERS = ["GPT-5.5", "Gemini 3.1 Pro Preview"]

# ---------------------------------------------------------------------------
# Helpers (mirrored from create_model_comparison_files.py)
# ---------------------------------------------------------------------------

_HUMAN_KEYWORDS = ("human",)
_LLM_KEYWORDS = ("llm",)
_METRICS_KEYWORD = "metrics"
_KAPPA_KEYWORDS = ("kappa", "cohens", "cohen")
_INTER_KEYWORDS = ("inter",)


def _identify_sheets(wb: openpyxl.Workbook) -> dict[str, str | None]:
    roles: dict[str, str | None] = {
        "ground_truth": None,
        "human_grading": None,
        "llm_grading": None,
        "metrics": None,
        "inter_rater": None,
        "weighted_kappa": None,
    }
    for name in wb.sheetnames:
        lo = name.lower()
        if _METRICS_KEYWORD in lo:
            roles["metrics"] = name
        elif any(k in lo for k in _KAPPA_KEYWORDS):
            roles["weighted_kappa"] = name
        elif any(k in lo for k in _INTER_KEYWORDS):
            roles["inter_rater"] = name
        elif any(k in lo for k in _LLM_KEYWORDS):
            roles["llm_grading"] = name
        elif any(k in lo for k in _HUMAN_KEYWORDS):
            roles["human_grading"] = name
        else:
            roles["ground_truth"] = name
    return roles


def _replace_sheet_ref(formula: str, old: str, new: str) -> str:
    if not formula:
        return formula
    result = formula.replace(f"'{old}'!", f"'{new}'!")
    if " " not in old:
        result = result.replace(f"{old}!", f"'{new}'!" if " " in new else f"{new}!")
    return result


def _update_all_formulas(wb: openpyxl.Workbook, old: str, new: str) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _replace_sheet_ref(cell.value, old, new)


def _find_source_file(example_dir: Path) -> Optional[Path]:
    pattern = str(
        example_dir
        / "Grading"
        / "2 stage"
        / "**"
        / "*CombinedHumanGradingVsClaude4.5SonnetGrading_Claude4.5SonnetGeneration.xlsx"
    )
    candidates = [
        Path(f)
        for f in glob.glob(pattern, recursive=True)
        if not os.path.basename(f).startswith("~$")
    ]
    return sorted(candidates)[-1] if candidates else None


_FNAME_RE = re.compile(
    r"^(.+?)_Grading_(\d-stage)_(\d{4}-\d{2}-\d{2})_(\d+-examples)_CombinedHumanGradingVsClaude4\.5SonnetGrading_Claude4\.5SonnetGeneration\.xlsx$"
)


def _parse_source_name(name: str) -> tuple[str, str, str, str] | None:
    m = _FNAME_RE.match(name)
    if not m:
        return None
    return m.group(1), m.group(2), m.group(3), m.group(4)


# ---------------------------------------------------------------------------
# Core: create one grading-comparison workbook
# ---------------------------------------------------------------------------


def create_grading_comparison_workbook(
    source_path: Path,
    new_grader: str,
    output_date: str,
    output_dir: Path,
) -> Path | None:
    """
    Build a new workbook comparing *new_grader* auto-grading against Claude 4.5 Sonnet
    auto-grading, both applied to the same (fixed) Claude 4.5 Sonnet generation.

    Returns the path of the saved workbook, or None on failure.
    """
    parsed = _parse_source_name(source_path.name)
    if not parsed:
        print(f"  [WARN] Cannot parse file name: {source_path.name}")
        return None

    example_prefix, stage, _src_date, n_examples = parsed

    # ------------------------------------------------------------------
    # 1. Load workbook
    # ------------------------------------------------------------------
    try:
        wb = load_workbook(source_path)
    except Exception as exc:
        print(f"  [ERROR] Cannot open {source_path.name}: {exc}")
        return None

    roles = _identify_sheets(wb)
    old_human = roles["human_grading"]
    old_llm = roles["llm_grading"]

    if not old_human:
        print(f"  [WARN] Human-grading sheet not found in {source_path.name}")
        return None
    if not old_llm:
        print(f"  [WARN] LLM-grading sheet not found in {source_path.name}")
        return None

    # ------------------------------------------------------------------
    # 2. Determine new sheet names (≤31 chars for Excel)
    # ------------------------------------------------------------------
    # e.g. "GPT-5.5 Grading" (15 chars), "Gemini 3.1 Pro Preview Grading" (30 chars)
    new_grader_sheet = f"{new_grader} Grading"
    new_baseline_sheet = "Claude 4.5 Sonnet Grading"  # 25 chars — the fixed baseline

    assert len(new_grader_sheet) <= 31, f"Sheet name too long: {new_grader_sheet!r}"

    # ------------------------------------------------------------------
    # 3. Update all formula references BEFORE renaming
    # ------------------------------------------------------------------
    _update_all_formulas(wb, old_human, new_grader_sheet)
    _update_all_formulas(wb, old_llm, new_baseline_sheet)

    # ------------------------------------------------------------------
    # 4. Rename the two grading sheets
    # ------------------------------------------------------------------
    wb[old_human].title = new_grader_sheet
    wb[old_llm].title = new_baseline_sheet

    # ------------------------------------------------------------------
    # 5. Prepare the new grader sheet (template — grade/notes cleared)
    #    Source: the old human-grading sheet keeps the element list intact,
    #    which matches the LLM sheet row-for-row. Only grades are cleared.
    # ------------------------------------------------------------------
    ws_new = wb[new_grader_sheet]

    # Update column C header
    header_row = next(ws_new.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        if isinstance(cell.value, str) and cell.value.lower() in (
            "human grading",
            "human",
            "consolidated human grading",
            "human grader",
        ):
            cell.value = f"{new_grader} Grading"

    # Clear grade (col C) and notes (col D) — placeholder for future grader output
    for row in ws_new.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.value = None
    for row in ws_new.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.value = None

    # ------------------------------------------------------------------
    # 6. Update column C header in the baseline (Claude 4.5 Sonnet) sheet
    # ------------------------------------------------------------------
    ws_baseline = wb[new_baseline_sheet]
    baseline_header = next(ws_baseline.iter_rows(min_row=1, max_row=1))
    for cell in baseline_header:
        if isinstance(cell.value, str) and cell.value.lower() in ("llm", "llm grading"):
            cell.value = "Claude 4.5 Sonnet Grading"

    # ------------------------------------------------------------------
    # 7. Metrics sheet: table titles + column J metadata
    #    A1  = "Graded by {new_grader}"
    #    A12 = "Graded by Claude 4.5 Sonnet"
    #    J4='Generation LLM', J5='Claude 4.5 Sonnet'
    #    J7='Grading LLMs',   J8={new_grader}, J9='Claude 4.5 Sonnet'
    # ------------------------------------------------------------------
    if roles["metrics"]:
        ws_metrics = wb[roles["metrics"]]
        ws_metrics["A1"] = f"Graded by {new_grader}"
        ws_metrics["A12"] = f"Graded by {GENERATION_LLM_DISPLAY}"
        for row_idx in range(4, 12):
            ws_metrics.cell(row=row_idx, column=10).value = None
        ws_metrics.cell(row=4, column=10).value = "Generation LLM"
        ws_metrics.cell(row=5, column=10).value = GENERATION_LLM_DISPLAY
        ws_metrics.cell(row=7, column=10).value = "Grading LLMs"
        ws_metrics.cell(row=8, column=10).value = new_grader
        ws_metrics.cell(row=9, column=10).value = GENERATION_LLM_DISPLAY

    # ------------------------------------------------------------------
    # 8. Inter-rater and Weighted Cohens Kappa: update column headers
    # ------------------------------------------------------------------
    for role_key in ("inter_rater", "weighted_kappa"):
        sheet_name = roles[role_key]
        if not sheet_name:
            continue
        ws = wb[sheet_name]
        hdr = next(ws.iter_rows(min_row=1, max_row=1))
        for cell in hdr:
            if isinstance(cell.value, str):
                if cell.value.lower() == "human":
                    cell.value = new_grader
                elif cell.value.lower() == "llm":
                    cell.value = GENERATION_LLM_DISPLAY

    # ------------------------------------------------------------------
    # 9. Construct output file name
    # ------------------------------------------------------------------
    grader_safe = new_grader.replace(
        " ", ""
    )  # "GPT-5.5" → "GPT-5.5", "Gemini 3.1 Pro Preview" → "Gemini3.1ProPreview"
    out_name = (
        f"{example_prefix}_Grading_{stage}_{output_date}"
        f"_{n_examples}_{grader_safe}AutoGradingVSClaude4.5SonnetAutoGrading"
        f"-Claude4.5SonnetGeneration.xlsx"
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / out_name
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--graders",
        nargs="+",
        default=DEFAULT_GRADERS,
        metavar="GRADER",
        help="Auto-grader model(s) to create comparison files for (default: %(default)s)",
    )
    parser.add_argument(
        "--date",
        default="2026-05-04",
        help="Date string to embed in the output file names (default: %(default)s)",
    )
    parser.add_argument(
        "--stage",
        default="2",
        choices=["1", "2"],
        help="Grading stage to process (default: %(default)s)",
    )
    args = parser.parse_args()

    example_dirs = sorted(
        d for d in BASE_DIR.iterdir() if d.is_dir() and (d / "Grading").is_dir()
    )

    if not example_dirs:
        print("[ERROR] No example directories found under", BASE_DIR)
        return

    print(f"Found {len(example_dirs)} example(s): {[d.name for d in example_dirs]}")
    print(f"Graders to compare against Claude 4.5 Sonnet: {args.graders}")
    print(f"Output date: {args.date}")
    print(f"Output folder: date-level folder (alongside other xlsx files)\n")

    for example_dir in example_dirs:
        source = _find_source_file(example_dir)
        if source is None:
            print(
                f"[SKIP] {example_dir.name}: "
                "no 2-stage CombinedHumanGradingVsClaude4.5SonnetGrading_Claude4.5SonnetGeneration.xlsx found"
            )
            continue

        print(f"[{example_dir.name}]  source: {source.name}")
        output_dir = (
            source.parent
        )  # same level as CombinedHumanGradingVsClaude4.5SonnetGrading_Claude4.5SonnetGeneration.xlsx

        for grader in args.graders:
            out = create_grading_comparison_workbook(
                source_path=source,
                new_grader=grader,
                output_date=args.date,
                output_dir=output_dir,
            )
            if out:
                print(f"    ✓  {out.name}")
            else:
                print(f"    ✗  Failed for grader: {grader}")

    print("\nDone.")


if __name__ == "__main__":
    main()
