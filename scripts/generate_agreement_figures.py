#!/usr/bin/env python3
"""
Generate confusion matrix figures from evaluation grading xlsx files.

Two modes are supported (--mode flag):

  human-vs-llm  (default)
    Compares Human grader scores vs Claude 4.5 Sonnet LLM auto-grader scores.
    Source files : *CombinedHumanGradingVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration.xlsx  (at date-folder level)
    Per-file output : <date_dir>/ConfusionMatrices_HumanVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration/
    Aggregated output : Global Analysis/ConfusionMatrices_HumanVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration/

  grader-vs-grader
    Compares a new LLM grader against the Claude 4.5 Sonnet baseline grader.
    Source files : *AutoGradingVSClaude4.5SonnetAutoGrading.xlsx  (at date-folder level)
    Per-file output : <date_dir>/ConfusionMatrices_NewGraderVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration/
    Aggregated output : Global Analysis/ConfusionMatrices_NewGraderVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration/

Usage:
  python generate_agreement_figures.py                         # human-vs-llm, all files
  python generate_agreement_figures.py --stage 2               # 2-stage files only
  python generate_agreement_figures.py --date 2026-03-16
  python generate_agreement_figures.py --mode grader-vs-grader
  python generate_agreement_figures.py --mode grader-vs-grader --stage 2
"""

import argparse
import glob
import os
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SCORE_LABELS = ["0", "0.5", "1"]
SCORE_TO_IDX = {0.0: 0, 0.5: 1, 1.0: 2}

# Ordered list of (display title, filename stem, optional type-filter string)
SCOPES: list[tuple[str, str, str | None]] = [
    ("Global", "global", None),
    ("State", "state", "State"),
    ("Transition", "transition", "Transition"),
    ("Composite State", "composite_state", "Composite State"),
    ("Guard", "guard", "Guard"),
    ("Action", "action", "Action"),
    ("History State", "history_state", "History State"),
    ("Region", "region", "Region"),
]

BASE_DIR = Path(__file__).parent.parent

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def find_xlsx_files(
    base_dir: Path,
    stage_filter: str | None = None,
    date_filter: str | None = None,
    file_pattern: str = "*CombinedHumanGradingVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration.xlsx",
) -> list[Path]:
    """Return xlsx files matching *file_pattern* under base_dir (no temp ~$ files)."""
    pattern = str(base_dir / "**" / file_pattern)
    files = [
        Path(f)
        for f in glob.glob(pattern, recursive=True)
        if not os.path.basename(f).startswith("~$")
    ]
    if stage_filter:
        tag = f"{stage_filter} stage"
        files = [f for f in files if tag in str(f)]
    if date_filter:
        files = [f for f in files if date_filter in str(f)]
    return sorted(files)


def _find_kappa_sheet(wb: openpyxl.Workbook) -> str | None:
    """Return the name of the weighted Cohen's kappa sheet, regardless of exact capitalisation."""
    for name in wb.sheetnames:
        lower = name.lower()
        if "kappa" in lower or ("cohen" in lower and "weight" in lower):
            return name
    return None


def load_grading_data(xlsx_path: Path) -> pd.DataFrame | None:
    """
    Load Human and LLM grades from the weighted Cohen's kappa sheet.
    Returns a DataFrame with columns: type, element, human, llm, is_additional.

    Regular rows have is_additional=False and human/llm are 0/0.5/1 grading scores.
    'additional elements' rows have is_additional=True and human/llm are raw *counts*
    of false-positive elements; they are handled separately in build_confusion_matrix.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"  [WARN] Could not open {xlsx_path.name}: {e}")
        return None

    sheet_name = _find_kappa_sheet(wb)
    if sheet_name is None:
        print(f"  [WARN] No weighted Cohen's kappa sheet found in {xlsx_path.name}")
        return None

    ws = wb[sheet_name]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        type_, element, human, llm = row[0], row[1], row[2], row[3]
        if type_ is None or element is None:
            break
        if human is None or llm is None:
            continue
        is_additional = str(element).strip().lower() == "additional elements"
        try:
            records.append(
                {
                    "type": str(type_).strip(),
                    "element": str(element).strip(),
                    "human": float(human),
                    "llm": float(llm),
                    "is_additional": is_additional,
                }
            )
        except (ValueError, TypeError):
            continue

    return pd.DataFrame(records) if records else None


def load_all(files: list[Path]) -> pd.DataFrame:
    """Load and concatenate grading data from all xlsx files."""
    frames = []
    for f in files:
        df = load_grading_data(f)
        if df is not None and len(df) > 0:
            df["source"] = f.name
            frames.append(df)
            print(f"  Loaded {len(df):3d} rows  ← {f.name[:75]}")
    if not frames:
        raise ValueError("No valid data found in any xlsx file.")
    return pd.concat(frames, ignore_index=True)


# ---------------------------------------------------------------------------
# Confusion matrix helpers
# ---------------------------------------------------------------------------


def build_confusion_matrix(
    df: pd.DataFrame, type_filter: str | None = None
) -> np.ndarray:
    """Build a 3×3 count matrix (rows = Human score, cols = LLM score).

    Regular rows (is_additional=False) contribute via their 0/0.5/1 scores.

    'Additional elements' rows carry raw counts of false-positive elements and
    are distributed across cells using the same MAX/MIN logic as the Excel sheet:
      - MIN(human, llm)          → [human=1, llm=1]  both graders agreed on extras
      - MAX(llm - human, 0)      → [human=0, llm=1]  LLM hallucinated extra elements
      - MAX(human - llm, 0)      → [human=1, llm=0]  human found extras LLM missed
    """
    if type_filter:
        df = df[df["type"] == type_filter]

    regular = df[~df["is_additional"]]
    additional = df[df["is_additional"]]

    matrix = np.zeros((3, 3), dtype=int)

    for _, row in regular.iterrows():
        h = SCORE_TO_IDX.get(row["human"])
        l = SCORE_TO_IDX.get(row["llm"])
        if h is not None and l is not None:
            matrix[h, l] += 1

    for _, row in additional.iterrows():
        h_count = int(row["human"])
        l_count = int(row["llm"])
        if h_count == 0 and l_count == 0:
            matrix[0, 0] += 1  # both agree there are no extra elements
        else:
            matrix[2, 2] += min(h_count, l_count)  # both agreed on extras
            matrix[0, 2] += max(l_count - h_count, 0)  # LLM found more extras
            matrix[2, 0] += max(h_count - l_count, 0)  # human found more extras

    return matrix


def row_normalize(matrix: np.ndarray) -> np.ndarray:
    """Normalize each row to [0, 1]; rows that sum to zero become NaN."""
    row_sums = matrix.sum(axis=1, keepdims=True).astype(float)
    row_sums[row_sums == 0] = np.nan
    return matrix / row_sums


# ---------------------------------------------------------------------------
# Core plotting primitive
# ---------------------------------------------------------------------------


def plot_confusion_heatmap(
    ax: plt.Axes,
    matrix: np.ndarray,
    title: str,
    show_ylabel: bool = True,
    show_xlabel: bool = True,
    x_label: str = "LLM Score",
    y_label: str = "Human Score",
) -> None:
    """Draw a single row-normalized confusion-matrix heatmap onto ax."""
    norm = row_normalize(matrix)

    annots = np.empty_like(norm, dtype=object)
    for i in range(3):
        for j in range(3):
            pct = norm[i, j]
            cnt = matrix[i, j]
            annots[i, j] = "—" if np.isnan(pct) else f"{pct:.0%}\n({cnt})"

    sns.heatmap(
        np.where(np.isnan(norm), 0, norm),
        annot=annots,
        fmt="",
        cmap="Blues",
        xticklabels=SCORE_LABELS,
        yticklabels=SCORE_LABELS,
        ax=ax,
        vmin=0,
        vmax=1,
        cbar=False,
        linewidths=0.5,
        linecolor="white",
    )
    for i in range(3):
        ax.add_patch(
            plt.Rectangle(
                (i, i), 1, 1, fill=False, edgecolor="black", lw=1.5, clip_on=False
            )
        )

    n_total = int(matrix.sum())
    n_agree = int(np.trace(matrix))
    agree_pct = n_agree / n_total if n_total > 0 else 0
    ax.set_title(f"{title}\n(agree {agree_pct:.0%}, n={n_total})", fontsize=9, pad=4)
    ax.set_xlabel(x_label if show_xlabel else "", fontsize=8)
    ax.set_ylabel(y_label if show_ylabel else "", fontsize=8)
    ax.tick_params(labelsize=8)


# ---------------------------------------------------------------------------
# Figure A: Combined 2×4 grid (unchanged from original)
# ---------------------------------------------------------------------------


def save_confusion_grid(
    df: pd.DataFrame,
    out_path: Path,
    title_prefix: str = "",
    comparison_title: str = "LLM vs Human Grader",
    x_label: str = "LLM Score",
    y_label: str = "Human Score",
) -> None:
    """Save the 2×4 combined confusion-matrix grid to out_path."""
    fig, axes = plt.subplots(2, 4, figsize=(14, 7))
    axes = axes.flatten()

    for idx, (title, _, type_filter) in enumerate(SCOPES):
        matrix = build_confusion_matrix(df, type_filter=type_filter)
        plot_confusion_heatmap(
            axes[idx],
            matrix,
            title,
            show_ylabel=(idx % 4 == 0),
            show_xlabel=True,
            x_label=x_label,
            y_label=y_label,
        )

    # No unused axes: 8 scopes fill the 2×4 grid exactly.
    sup = f"{title_prefix} — " if title_prefix else ""
    fig.suptitle(
        f"{sup}{comparison_title}: Row-Normalized Confusion Matrices\n"
        f"(rows = {y_label.replace(' Score', '')} / cols = {x_label.replace(' Score', '')})",
        fontsize=10,
        y=1.01,
    )
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved → {out_path.relative_to(BASE_DIR)}")


# ---------------------------------------------------------------------------
# Figure B: 8 individual confusion matrices
# ---------------------------------------------------------------------------


def save_individual_confusion_matrices(
    df: pd.DataFrame,
    out_dir: Path,
    title_prefix: str = "",
    file_prefix: str = "",
    comparison_title: str = "LLM vs Human",
    x_label: str = "LLM Score",
    y_label: str = "Human Score",
) -> None:
    """Save one figure per scope (global.png, state.png, …) into out_dir."""
    out_dir.mkdir(parents=True, exist_ok=True)
    sup = f"{title_prefix} — " if title_prefix else ""

    for title, stem, type_filter in SCOPES:
        matrix = build_confusion_matrix(df, type_filter=type_filter)
        fig, ax = plt.subplots(figsize=(4, 3.5))
        plot_confusion_heatmap(
            ax,
            matrix,
            title,
            show_ylabel=True,
            show_xlabel=True,
            x_label=x_label,
            y_label=y_label,
        )
        fig.suptitle(
            f"{sup}{comparison_title}: {title}",
            fontsize=9,
            y=1.02,
        )
        fig.tight_layout()
        fname = f"{file_prefix}_{stem}.png" if file_prefix else f"{stem}.png"
        out_path = out_dir / fname
        fig.savefig(out_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        print(f"  Saved → {out_path.relative_to(BASE_DIR)}")


# ---------------------------------------------------------------------------
# High-level: write all confusion matrix outputs into a given directory
# ---------------------------------------------------------------------------


def generate_confusion_outputs(
    df: pd.DataFrame,
    cm_dir: Path,
    title_prefix: str = "",
    file_prefix: str = "",
    comparison_title: str = "LLM vs Human Grader",
    x_label: str = "LLM Score",
    y_label: str = "Human Score",
) -> None:
    """
    Write into cm_dir:
      [file_prefix_]confusion_matrices.png   (combined 2×4 grid)
      [file_prefix_]global.png, state.png, … (8 individual figures)
    """
    cm_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n  → {cm_dir.relative_to(BASE_DIR)}  ({len(df)} rows)")
    grid_name = (
        f"{file_prefix}_confusion_matrices.png"
        if file_prefix
        else "confusion_matrices.png"
    )
    save_confusion_grid(
        df,
        cm_dir / grid_name,
        title_prefix,
        comparison_title=comparison_title,
        x_label=x_label,
        y_label=y_label,
    )
    save_individual_confusion_matrices(
        df,
        cm_dir,
        title_prefix,
        file_prefix=file_prefix,
        comparison_title=comparison_title,
        x_label=x_label,
        y_label=y_label,
    )


# ---------------------------------------------------------------------------
# Label helper
# ---------------------------------------------------------------------------


def label_from_path(xlsx_path: Path) -> str:
    """
    Derive a human-readable label from the xlsx path, e.g.:
      Printer — 2 stage — 2026-03-16
    Works for files at date-folder level AND inside subfolders.
    Expected layout: …/<Project>/Grading/<N stage>/<Date>/[<Subfolder>/]<file>.xlsx
    """
    parts = xlsx_path.parts
    try:
        # Find the 'Grading' directory (not a subfolder containing 'Grading')
        grading_idx = parts.index("Grading")
        project = parts[grading_idx - 1]  # e.g. "Printer"
        stage_dir = parts[grading_idx + 1]  # e.g. "2 stage"
        date_dir = parts[grading_idx + 2]  # e.g. "2026-03-16"
    except (ValueError, IndexError):
        # Fallback: use grandparent or parent names
        date_dir = xlsx_path.parent.name
        stage_dir = xlsx_path.parent.parent.name
        project = xlsx_path.parent.parent.parent.name
    return f"{project} — {stage_dir} — {date_dir}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate confusion matrix figures from evaluation grading xlsx files."
    )
    parser.add_argument(
        "--mode",
        choices=["human-vs-llm", "grader-vs-grader"],
        default="human-vs-llm",
        help=(
            "human-vs-llm (default): Human grader vs Claude 4.5 Sonnet LLM auto-grader "
            "from *CombinedHumanGradingVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration.xlsx files. "
            "grader-vs-grader: New LLM auto-grader vs Claude 4.5 Sonnet baseline grader (fixed Claude 4.5 Sonnet generation), "
            "from *AutoGradingVSClaude4.5SonnetAutoGrading.xlsx files (at date-folder level)."
        ),
    )
    parser.add_argument(
        "--stage",
        choices=["1", "2"],
        default=None,
        help="Restrict to '1 stage' or '2 stage' files. Default: all.",
    )
    parser.add_argument(
        "--date",
        default=None,
        help="Restrict to files whose path contains this date string (e.g. 2026-03-16).",
    )
    args = parser.parse_args()

    # ── Mode-specific settings ─────────────────────────────────────────
    if args.mode == "grader-vs-grader":
        file_pattern = "*AutoGradingVSClaude4.5SonnetAutoGrading.xlsx"
        comparison_title = "New Grader vs Claude 4.5 Sonnet Grader"
        x_label = "Claude 4.5 Sonnet Score"
        y_label = "New Grader Score"
        per_file_cm_subdir = "ConfusionMatrices_NewGraderVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration"
        global_cm_dir = (
            BASE_DIR
            / "Global Analysis"
            / "ConfusionMatrices_NewGraderVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration"
        )
        global_prefix = "AllExamples_Grading_NewGraderVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration"
    else:  # human-vs-llm
        file_pattern = "*CombinedHumanGradingVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration.xlsx"
        comparison_title = "LLM vs Human Grader"
        x_label = "LLM Score"
        y_label = "Human Score"
        per_file_cm_subdir = "ConfusionMatrices_HumanVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration"
        global_cm_dir = (
            BASE_DIR
            / "Global Analysis"
            / "ConfusionMatrices_HumanVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration"
        )
        global_prefix = "AllExamples_Grading_CombinedHumanGradingVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration"

    print(
        f"Mode: {args.mode} | "
        f"stage={args.stage or 'any'} | date={args.date or 'any'}"
    )
    files = find_xlsx_files(
        BASE_DIR,
        stage_filter=args.stage,
        date_filter=args.date,
        file_pattern=file_pattern,
    )
    print(f"Found {len(files)} file(s).\n")

    if not files:
        print("No files matched. Exiting.")
        return

    # ------------------------------------------------------------------
    # 1. Per-file: output into <date_dir>/ConfusionMatrices_HumanVsClaude4.5SonnetAutoGrading_Claude4.5SonnetGeneration/
    #    (file names follow the xlsx stem, which now includes the full grading type)
    # ------------------------------------------------------------------
    print("=" * 60)
    print("Per-file confusion matrices")
    print("=" * 60)
    all_frames: list[pd.DataFrame] = []
    for xlsx_path in files:
        df = load_grading_data(xlsx_path)
        if df is None or len(df) == 0:
            print(f"  [SKIP] {xlsx_path.name}")
            continue
        all_frames.append(df)
        cm_dir = xlsx_path.parent / per_file_cm_subdir
        label = label_from_path(xlsx_path)
        generate_confusion_outputs(
            df,
            cm_dir,
            title_prefix=label,
            file_prefix=xlsx_path.stem,
            comparison_title=comparison_title,
            x_label=x_label,
            y_label=y_label,
        )

    if not all_frames:
        print("No valid data found. Exiting.")
        return

    # ------------------------------------------------------------------
    # 2. Aggregate: output into the mode-specific global directory
    # ------------------------------------------------------------------
    print()
    print("=" * 60)
    print("Aggregated confusion matrices (all files)")
    print("=" * 60)
    df_all = pd.concat(all_frames, ignore_index=True)
    generate_confusion_outputs(
        df_all,
        global_cm_dir,
        title_prefix="All Examples",
        file_prefix=global_prefix,
        comparison_title=comparison_title,
        x_label=x_label,
        y_label=y_label,
    )

    print("\nDone.")


if __name__ == "__main__":
    main()
