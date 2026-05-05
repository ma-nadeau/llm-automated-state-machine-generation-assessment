#!/usr/bin/env python3
"""
Create model-comparison Excel templates from existing CombinedHumanGradingVsClaude4.5SonnetGrading_Claude4.5SonnetGeneration.xlsx files.

For each benchmark example, this script produces new workbooks that compare a NEW generation
model (e.g. GPT-5.5, Gemini 3.1 Pro Preview) against Claude 4.5 Sonnet, where both outputs are graded by
the same fixed auto-grader (Claude 4.5 Sonnet).

Experiment type: GENERATION varies, GRADER fixed (Claude 4.5 Sonnet).
See create_grading_comparison_files.py for the inverse experiment (generation fixed, grader varies).

Output naming convention:
  {Example}_Grading_2-stage_{date}_{N}-examples_
    {NewModel}GenerationVSClaude4.5SonnetGeneration_Claude4.5SonnetAutoGrading.xlsx

Output path: {date_dir}/{filename}.xlsx  (same level as the other xlsx files)

Changes made inside each new workbook:
  • Ground-truth sheet  – unchanged
  • Human-grading sheet – renamed to "{New Model} Generation";
                          grade column (C) cleared (placeholder for future data);
                          column C header updated
  • LLM-grading sheet  – renamed to "Claude 4.5 Sonnet Gen";
                          column C header updated
  • Metrics            – column J rewritten: Grading LLM + Generation LLMs
  • Inter-rater        – column C/D headers updated
  • Weighted Cohens Kappa – column C/D headers updated
  • All formula strings – old sheet-name references replaced with new names

Usage:
  # Default: creates GPT-5.5 and Gemini 3.1 Pro Preview comparison files for all 9 examples
  python create_model_comparison_files.py

  # Specific model(s) only
  python create_model_comparison_files.py --models "GPT-5.5"
  python create_model_comparison_files.py --models "GPT-5.5" "Gemini 3.1 Pro Preview"

  # 2026-05-04 is used as the output date by default; override with --date
  python create_model_comparison_files.py --date 2026-05-04
"""

from __future__ import annotations

import argparse
import glob
import os
import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent

# The fixed grader used in all evaluations
GRADER_DISPLAY = "Claude 4.5 Sonnet"  # human-readable
GRADER_FILE = "Claude4.5Sonnet"  # used in file/sheet names (no spaces)

# New LLM sheet name (replacing "LLM Grading"). ≤31 chars for Excel.
NEW_LLM_SHEET_LABEL = "Claude 4.5 Sonnet Gen"  # 21 chars

# Column C header in the new LLM grading sheet
NEW_LLM_COL_HEADER = "Claude 4.5 Sonnet Generation"

# Default models to create comparison files for
DEFAULT_MODELS = ["GPT-5.5", "Gemini 3.1 Pro Preview"]

# ---------------------------------------------------------------------------
# Helper: identify sheet roles inside a workbook
# ---------------------------------------------------------------------------

# Keywords that signal each sheet role (checked against lowercased sheet name)
_HUMAN_KEYWORDS = (
    "human",
)  # present in human-grading sheet, absent in ground-truth/LLM
_LLM_KEYWORDS = ("llm",)
_METRICS_KEYWORD = "metrics"
_KAPPA_KEYWORDS = ("kappa", "cohens", "cohen")
_INTER_KEYWORDS = ("inter",)


def _identify_sheets(wb: openpyxl.Workbook) -> dict[str, str | None]:
    """Return a mapping of role → sheet name for the six expected sheet roles."""
    roles: dict[str, str | None] = {
        "ground_truth": None,
        "human_grading": None,
        "llm_grading": None,
        "metrics": None,
        "inter_rater": None,
        "weighted_kappa": None,
    }
    for name in wb.sheetnames:
        lo = name.lower()
        if _METRICS_KEYWORD in lo:
            roles["metrics"] = name
        elif any(k in lo for k in _KAPPA_KEYWORDS):
            roles["weighted_kappa"] = name
        elif any(k in lo for k in _INTER_KEYWORDS):
            roles["inter_rater"] = name
        elif any(k in lo for k in _LLM_KEYWORDS):
            roles["llm_grading"] = name
        elif any(k in lo for k in _HUMAN_KEYWORDS):
            roles["human_grading"] = name
        else:
            # Whatever's left is the ground-truth sheet
            roles["ground_truth"] = name
    return roles


# ---------------------------------------------------------------------------
# Helper: update formula string when a sheet is renamed
# ---------------------------------------------------------------------------


def _replace_sheet_ref(formula: str, old: str, new: str) -> str:
    """Replace 'OldSheet'! and OldSheet! references inside a formula string."""
    if not formula:
        return formula
    # Quoted form first (handles names with spaces)
    result = formula.replace(f"'{old}'!", f"'{new}'!")
    # Unquoted form (only if old name has no spaces)
    if " " not in old:
        result = result.replace(f"{old}!", f"'{new}'!" if " " in new else f"{new}!")
    return result


def _update_all_formulas(wb: openpyxl.Workbook, old: str, new: str) -> None:
    """Walk every cell in every sheet and patch sheet-name references in formulas."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _replace_sheet_ref(cell.value, old, new)


# ---------------------------------------------------------------------------
# Helper: find the most recent 2-stage source file for an example folder
# ---------------------------------------------------------------------------


def _find_source_file(example_dir: Path) -> Optional[Path]:
    pattern = str(
        example_dir
        / "Grading"
        / "2 stage"
        / "**"
        / "*CombinedHumanGradingVsClaude4.5SonnetGrading_Claude4.5SonnetGeneration.xlsx"
    )
    candidates = [
        Path(f)
        for f in glob.glob(pattern, recursive=True)
        if not os.path.basename(f).startswith("~$")
    ]
    return sorted(candidates)[-1] if candidates else None


# ---------------------------------------------------------------------------
# Helper: extract metadata from the source file name
# ---------------------------------------------------------------------------

_FNAME_RE = re.compile(
    r"^(.+?)_Grading_(\d-stage)_(\d{4}-\d{2}-\d{2})_(\d+-examples)_CombinedHumanGradingVsClaude4\.5SonnetGrading_Claude4\.5SonnetGeneration\.xlsx$"
)


def _parse_source_name(name: str) -> tuple[str, str, str, str] | None:
    """Return (example_prefix, stage, date, n_examples) or None."""
    m = _FNAME_RE.match(name)
    if not m:
        return None
    return m.group(1), m.group(2), m.group(3), m.group(4)


# ---------------------------------------------------------------------------
# Core: create one model-comparison workbook
# ---------------------------------------------------------------------------


def create_comparison_workbook(
    source_path: Path,
    new_model: str,
    output_date: str,
    output_dir: Path,
) -> Path | None:
    """
    Build a new workbook comparing *new_model* generation against Claude 4.5 Sonnet
    generation, both auto-graded by Claude 4.5 Sonnet.

    Returns the path of the saved workbook, or None on failure.
    """
    parsed = _parse_source_name(source_path.name)
    if not parsed:
        print(f"  [WARN] Cannot parse file name: {source_path.name}")
        return None

    example_prefix, stage, _src_date, n_examples = parsed

    # ------------------------------------------------------------------
    # 1. Load workbook
    # ------------------------------------------------------------------
    try:
        wb = load_workbook(source_path)
    except Exception as exc:
        print(f"  [ERROR] Cannot open {source_path.name}: {exc}")
        return None

    roles = _identify_sheets(wb)

    old_human = roles["human_grading"]
    old_llm = roles["llm_grading"]

    if not old_human:
        print(f"  [WARN] Human-grading sheet not found in {source_path.name}")
        return None
    if not old_llm:
        print(f"  [WARN] LLM-grading sheet not found in {source_path.name}")
        return None

    # ------------------------------------------------------------------
    # 2. Determine new sheet names (must be ≤31 chars)
    # ------------------------------------------------------------------
    # e.g. "GPT-5.5 Generation" (18 chars), "Gemini 3.1 Pro Preview Generation" (34 chars)
    new_human_sheet = f"{new_model} Generation"
    assert len(new_human_sheet) <= 31, f"Sheet name too long: {new_human_sheet}"

    new_llm_sheet = NEW_LLM_SHEET_LABEL  # "Claude 4.5 Sonnet Gen" (21 chars)

    # ------------------------------------------------------------------
    # 3. Update all formula references BEFORE renaming
    # ------------------------------------------------------------------
    _update_all_formulas(wb, old_human, new_human_sheet)
    _update_all_formulas(wb, old_llm, new_llm_sheet)

    # ------------------------------------------------------------------
    # 4. Rename the two data sheets
    # ------------------------------------------------------------------
    wb[old_human].title = new_human_sheet
    wb[old_llm].title = new_llm_sheet

    # ------------------------------------------------------------------
    # 5. Update column C header in the (renamed) human sheet and clear data
    # ------------------------------------------------------------------
    ws_human = wb[new_human_sheet]
    # Row 1: update the grade-column header (column C = index 2 in 0-based)
    header_row = next(ws_human.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        if isinstance(cell.value, str) and cell.value.lower() in (
            "human grading",
            "human",
            "consolidated human grading",
            "human grader",
        ):
            cell.value = f"{new_model} Generation"

    # Clear grade data in column C (rows 2 onwards) – it is a placeholder for new LLM data
    for row in ws_human.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.value = None

    # Also clear column D (Notes) data so no human-grading notes bleed through
    for row in ws_human.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.value = None

    # ------------------------------------------------------------------
    # 6. Update column C header in the (renamed) LLM sheet
    # ------------------------------------------------------------------
    ws_llm = wb[new_llm_sheet]
    llm_header_row = next(ws_llm.iter_rows(min_row=1, max_row=1))
    for cell in llm_header_row:
        if isinstance(cell.value, str) and cell.value.lower() in ("llm", "llm grading"):
            cell.value = NEW_LLM_COL_HEADER

    # ------------------------------------------------------------------
    # 7. Update Metrics sheet column J metadata + table title rows
    #    Layout: J1=Date, J2=<date value>,
    #            J4='Grading LLM', J5=GRADER_DISPLAY,
    #            J7='Generation LLMs', J8=new_model, J9=GRADER_DISPLAY
    #    Clear legacy Grader ID / LLM Used rows (rows 4-11, then rewrite).
    #    Table titles: A1='Generated by {new_model}', A12='Generated by {GRADER}'
    # ------------------------------------------------------------------
    if roles["metrics"]:
        ws_metrics = wb[roles["metrics"]]
        # Table titles
        ws_metrics["A1"] = f"Generated by {new_model}"
        ws_metrics["A12"] = f"Generated by {GRADER_DISPLAY}"
        # Clear rows 4-11 of column J (column index 10, 1-based)
        for row_idx in range(4, 12):
            ws_metrics.cell(row=row_idx, column=10).value = None
        # Write new metadata block
        ws_metrics.cell(row=4, column=10).value = "Grading LLM"
        ws_metrics.cell(row=5, column=10).value = GRADER_DISPLAY
        ws_metrics.cell(row=7, column=10).value = "Generation LLMs"
        ws_metrics.cell(row=8, column=10).value = new_model
        ws_metrics.cell(row=9, column=10).value = GRADER_DISPLAY

    # ------------------------------------------------------------------
    # 8. Update Inter-rater and Weighted Cohens Kappa column C/D headers
    # ------------------------------------------------------------------
    for role_key in ("inter_rater", "weighted_kappa"):
        sheet_name = roles[role_key]
        if not sheet_name:
            continue
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        for cell in header_row:
            if isinstance(cell.value, str):
                if cell.value.lower() == "human":
                    cell.value = new_model
                elif cell.value.lower() == "llm":
                    cell.value = GRADER_DISPLAY  # "Claude 4.5 Sonnet"

    # ------------------------------------------------------------------
    # 9. Construct output file name following the spec
    # ------------------------------------------------------------------
    # Spec: {prefix}GenerationVSClaude4.5SonnetGeneration_Claude4.5SonnetAutoGrading
    model_safe = new_model.replace(" ", "")  # e.g. "GPT-5.5", "Gemini3.1ProPreview"
    comparison_suffix = (
        f"{model_safe}GenerationVSClaude4.5SonnetGeneration"
        f"_Claude4.5SonnetAutoGrading"
    )
    out_name = (
        f"{example_prefix}_Grading_{stage}_{output_date}"
        f"_{n_examples}_{comparison_suffix}.xlsx"
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / out_name
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--models",
        nargs="+",
        default=DEFAULT_MODELS,
        metavar="MODEL",
        help="Generation model(s) to create comparison files for (default: %(default)s)",
    )
    parser.add_argument(
        "--date",
        default="2026-05-04",
        help="Date string to embed in the output file names (default: %(default)s)",
    )
    parser.add_argument(
        "--stage",
        default="2",
        choices=["1", "2"],
        help="Grading stage to process (default: %(default)s)",
    )
    args = parser.parse_args()

    # Discover all example directories (immediate children of BASE_DIR that
    # contain a Grading/ folder)
    example_dirs = sorted(
        d for d in BASE_DIR.iterdir() if d.is_dir() and (d / "Grading").is_dir()
    )

    if not example_dirs:
        print("[ERROR] No example directories found under", BASE_DIR)
        return

    print(f"Found {len(example_dirs)} example(s): {[d.name for d in example_dirs]}")
    print(f"Models to compare: {args.models}")
    print(f"Output date: {args.date}\n")

    for example_dir in example_dirs:
        source = _find_source_file(example_dir)
        if source is None:
            print(
                f"[SKIP] {example_dir.name}: no 2-stage CombinedHumanGradingVsClaude4.5SonnetGrading_Claude4.5SonnetGeneration.xlsx found"
            )
            continue

        print(f"[{example_dir.name}]  source: {source.name}")

        # Place new files alongside the other xlsx files at the date-folder level
        output_dir = source.parent

        for model in args.models:
            out = create_comparison_workbook(
                source_path=source,
                new_model=model,
                output_date=args.date,
                output_dir=output_dir,
            )
            if out:
                print(f"    ✓  {out.name}")
            else:
                print(f"    ✗  Failed for model: {model}")

    print("\nDone.")


if __name__ == "__main__":
    main()
